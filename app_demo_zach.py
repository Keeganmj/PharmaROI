# app.py
# PharmaROI Intelligence — V4 (Baseline vs Dario Comparison)
# Run: streamlit run app.py

from __future__ import annotations

import copy
from dataclasses import dataclass
from typing import List, Optional
import io
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import streamlit as st
import altair as alt

try:
    import pandas as pd
except Exception:
    pd = None

# =============================================================================
# COLOR PALETTE
# =============================================================================
COLORS = {
    "primary": "#0F6CBD",
    "revenue": "#0F6CBD",
    "costs": "#9CA3AF",
    "profit": "#10B981",
    "warning": "#F59E0B",
    "danger": "#EF4444",
    "muted": "#6B7280",
    "baseline": "#6B7280",      # Gray for baseline
    "dario": "#0F6CBD",         # Blue for Dario
    "incremental": "#10B981",   # Green for incremental/lift
}

TAB_PALETTE = [
    "#0F6CBD", "#10B981", "#F59E0B", "#EF4444",
    "#8B5CF6", "#EC4899", "#06B6D4", "#84CC16",
]

# =============================================================================
# FUNNEL STAGE DEFINITIONS
# =============================================================================
STAGE_NAMES: List[str] = [
    "Total Addressable Market for MASH",
    "F2 and F3",
    "MASH patients diagnosed",
    "Madrigal access to MASH patients",
    "Frequent users of online and social media resources",
    "Activation within 90 days onto Dario Connect for MASH",
    "Schedule telemedicine appointment",
    "Keep telemedicine appointment",
    "Obtain prescription for biopsy",
    "Get biopsy lab test",
    "Get positive lab results",
    "Complete post lab result consultation",
    "Get prescription for Rezdiffra",
]

NUM_STAGES = len(STAGE_NAMES)

# =============================================================================
# DEFAULT SCENARIO CONFIGURATIONS
# =============================================================================
def get_default_scenario(scenario_type: str = "dario") -> dict:
    """
    Returns default values for a scenario (baseline or dario).
    Baseline has lower conversion rates and different CAC assumptions.
    """
    if scenario_type == "baseline":
        return {
            "base_population": 10_000_000,
            "ratios": [1.00, 0.30, 0.12, 0.18, 0.60, 0.25, 0.10, 0.65, 0.85, 0.60, 0.80, 0.35, 0.80],
            "cac": [0.0, 0.0, 0.0, 0.0, 0.0, 8.0, 50.0, 65.0, 65.0, 90.0, 100.0, 200.0, 220.0],
            "arpp": 47_400.0,
            "treatment_years": 1.0,
            "discount": 0.68,
            "stage_active": [True] * NUM_STAGES,
            "platform_costs": {
                "dario_connect_config": 0.0,
                "dario_care_config": 0.0,
                "sub_dario_connect": 0.0,
                "sub_dario_care": 0.0,
                "maintenance_support": 0.0,
            },
            # CAC mode: "direct" or "sensitivity"
            "cac_mode": "direct",
            # Sensitivity settings (used when cac_mode == "sensitivity")
            "cac_sensitivity": {
                "min": 5.0,
                "max": 20.0,
                "step": 1.0,
                "base": 10.0,  # Base case for sensitivity
            },
        }
    else:  # dario
        return {
            "base_population": 10_000_000,
            "ratios": [1.00, 0.35, 0.16, 0.22, 0.75, 0.40, 0.15, 0.80, 1.00, 0.75, 0.90, 0.50, 0.90],
            "cac": [0.0, 0.0, 0.0, 0.0, 0.0, 10.0, 67.0, 83.0, 83.0, 111.0, 123.0, 247.0, 274.0],
            "arpp": 47_400.0,
            "treatment_years": 1.0,
            "discount": 0.68,
            "stage_active": [True] * NUM_STAGES,
            "platform_costs": {
                "dario_connect_config": 500_000.0,
                "dario_care_config": 500_000.0,
                "sub_dario_connect": 1_000_000.0,
                "sub_dario_care": 250_000.0,
                "maintenance_support": 250_000.0,
            },
            "cac_mode": "direct",
            "cac_sensitivity": {
                "min": 5.0,
                "max": 20.0,
                "step": 1.0,
                "base": 10.0,
            },
        }


def get_default_model() -> dict:
    """
    Returns the default structure for a model with shared settings,
    baseline scenario, and dario scenario.
    """
    return {
        "shared": {
            "stage_names": STAGE_NAMES[:],
            "use_shared_base_population": True,
            "shared_base_population": 10_000_000,
        },
        "baseline": get_default_scenario("baseline"),
        "dario": get_default_scenario("dario"),
    }


def get_zero_model() -> dict:
    """Returns a zeroed-out model for fresh starts."""
    zero_scenario = {
        "base_population": 0,
        "ratios": [0.0] * NUM_STAGES,
        "cac": [0.0] * NUM_STAGES,
        "arpp": 0.0,
        "treatment_years": 1.0,
        "discount": 0.0,
        "stage_active": [True] * NUM_STAGES,
        "platform_costs": {
            "dario_connect_config": 0.0,
            "dario_care_config": 0.0,
            "sub_dario_connect": 0.0,
            "sub_dario_care": 0.0,
            "maintenance_support": 0.0,
        },
        "cac_mode": "direct",
        "cac_sensitivity": {"min": 0.0, "max": 0.0, "step": 1.0, "base": 0.0},
    }
    return {
        "shared": {
            "stage_names": ["Stage " + str(i+1) for i in range(NUM_STAGES)],
            "use_shared_base_population": True,
            "shared_base_population": 0,
        },
        "baseline": copy.deepcopy(zero_scenario),
        "dario": copy.deepcopy(zero_scenario),
    }


# =============================================================================
# MIGRATION HELPER — Upgrade old models to new structure
# =============================================================================
def migrate_old_model(old_model: dict) -> dict:
    """
    Converts an old single-scenario model to the new baseline/dario structure.
    The old model becomes the 'dario' scenario; baseline gets defaults.
    """
    # Check if already in new format
    if "shared" in old_model and "baseline" in old_model and "dario" in old_model:
        return old_model
    
    # Create new structure
    new_model = get_default_model()
    
    # Migrate shared settings
    new_model["shared"]["stage_names"] = old_model.get("stage_names", STAGE_NAMES[:])
    new_model["shared"]["shared_base_population"] = old_model.get("base_population", 10_000_000)
    
    # Migrate old model to dario scenario
    new_model["dario"]["base_population"] = old_model.get("base_population", 10_000_000)
    new_model["dario"]["ratios"] = old_model.get("ratios", get_default_scenario("dario")["ratios"])
    new_model["dario"]["cac"] = old_model.get("cac", get_default_scenario("dario")["cac"])
    new_model["dario"]["arpp"] = old_model.get("arpp", 47_400.0)
    new_model["dario"]["treatment_years"] = old_model.get("treatment_years", 1.0)
    new_model["dario"]["discount"] = old_model.get("discount", 0.68)
    new_model["dario"]["stage_active"] = old_model.get("stage_active", [True] * NUM_STAGES)
    new_model["dario"]["platform_costs"] = old_model.get("platform_costs", get_default_scenario("dario")["platform_costs"])
    
    # Baseline gets slightly lower conversion rates (estimated)
    baseline_ratios = [r * 0.85 for r in new_model["dario"]["ratios"]]
    baseline_ratios[0] = 1.0  # First stage is always 1.0
    new_model["baseline"]["ratios"] = baseline_ratios
    new_model["baseline"]["base_population"] = new_model["dario"]["base_population"]
    new_model["baseline"]["arpp"] = new_model["dario"]["arpp"]
    new_model["baseline"]["treatment_years"] = new_model["dario"]["treatment_years"]
    new_model["baseline"]["discount"] = new_model["dario"]["discount"]
    
    return new_model


# =============================================================================
# FORMATTING HELPERS
# =============================================================================
def clamp(x, lo, hi):
    """Clamp a value between lo and hi."""
    return max(lo, min(hi, float(x)))


def money(x) -> str:
    """Format as currency."""
    if x != x:  # NaN check
        return "—"
    return f"${x:,.0f}"


def number(x) -> str:
    """Format as number with commas."""
    if x != x:
        return "—"
    return f"{x:,.0f}"


def pct(x) -> str:
    """Format as percentage."""
    if x != x:
        return "—"
    return f"{x*100:,.1f}%"


def roix(x) -> str:
    """Format as ROI multiplier."""
    if x != x:
        return "—"
    return f"{x:,.2f}x"


def delta_fmt(x, fmt_type="number") -> str:
    """Format a delta value with +/- sign."""
    if x != x:
        return "—"
    sign = "+" if x > 0 else ""
    if fmt_type == "money":
        return f"{sign}${x:,.0f}"
    elif fmt_type == "pct":
        return f"{sign}{x*100:,.1f}%"
    else:
        return f"{sign}{x:,.0f}"


# =============================================================================
# CORE DATA STRUCTURES
# =============================================================================
@dataclass(frozen=True)
class StageInput:
    """Input parameters for a single funnel stage."""
    name: str
    active: bool
    ratio: float
    cac: float


@dataclass(frozen=True)
class StageResult:
    """Computed results for a single funnel stage."""
    name: str
    active: bool
    ratio_used: float
    patients: float
    cac_per_patient: float
    stage_cac: float
    cumulative_cac: float


# =============================================================================
# CORE COMPUTATION FUNCTIONS
# =============================================================================
def build_stage_inputs(scenario: dict, stage_names: List[str]) -> List[StageInput]:
    """
    Build a list of StageInput objects from scenario data.
    
    Args:
        scenario: Dictionary containing ratios, cac, stage_active
        stage_names: List of stage names (from shared settings)
    
    Returns:
        List of StageInput objects
    """
    stages = []
    for idx, name in enumerate(stage_names):
        stages.append(StageInput(
            name=name,
            active=bool(scenario["stage_active"][idx]),
            ratio=float(scenario["ratios"][idx]) if idx > 0 else 1.0,
            cac=float(scenario["cac"][idx]),
        ))
    return stages


def compute_funnel(stages: List[StageInput], base_population: float) -> List[StageResult]:
    """
    Calculate funnel progression from base population through all stages.
    
    CAC Logic:
    - Stages 1-5 (idx 0-4): No CAC applied (top-of-funnel awareness)
    - Stage 6 (idx 5): CAC is entered per-patient, creates the CAC pool
    - Stages 7+ (idx 6+): CAC per patient is calculated from the Stage 6 pool
    
    Args:
        stages: List of StageInput objects
        base_population: Starting patient count
    
    Returns:
        List of StageResult objects with computed values
    """
    results = []
    prev_patients = max(0.0, float(base_population))
    total_cac_pool = 0.0

    for idx, s in enumerate(stages):
        # Calculate patient count for this stage
        if idx == 0:
            patients = prev_patients
            ratio_used = 1.0
        else:
            ratio_used = 1.0 if not s.active else clamp(s.ratio, 0.0, 1.0)
            patients = prev_patients * ratio_used

        # Calculate CAC based on stage position
        if idx < 5:
            # Stages 1-5: No CAC
            cac_pp = 0.0
            stage_cac = 0.0
            cumulative = 0.0
        elif idx == 5:
            # Stage 6: CAC per patient creates the pool
            cac_pp = 0.0 if not s.active else max(0.0, float(s.cac))
            stage_cac = patients * cac_pp
            total_cac_pool = stage_cac
            cumulative = total_cac_pool
        else:
            # Stages 7+: CAC distributed from pool
            cumulative = total_cac_pool
            cac_pp = (total_cac_pool / patients) if patients > 0 else 0.0
            stage_cac = cac_pp * patients

        results.append(StageResult(
            name=s.name,
            active=s.active,
            ratio_used=ratio_used,
            patients=patients,
            cac_per_patient=cac_pp,
            stage_cac=stage_cac,
            cumulative_cac=cumulative,
        ))
        prev_patients = patients

    return results


def compute_financials(
    treated_patients: float,
    arpp: float,
    treatment_years: float,
    discount: float,
    funnel_cac_total: float,
    platform_costs: float = 0.0
) -> dict:
    """
    Calculate financial metrics for a scenario.
    
    Formulas:
    - Gross Revenue = Treated Patients × ARPP × Treatment Years
    - Net Revenue = Gross Revenue × (1 - Discount)
    - Total Cost = Funnel CAC + Platform Costs
    - Net Profit = Net Revenue - Total Cost
    - ROI = Net Revenue / Total Cost
    
    Args:
        treated_patients: Number of patients completing the funnel
        arpp: Annual Revenue Per Patient
        treatment_years: Average years of treatment
        discount: Gross-to-net discount rate (0.68 = 68% discount)
        funnel_cac_total: Total customer acquisition cost from funnel
        platform_costs: Additional platform/technology costs
    
    Returns:
        Dictionary with all financial metrics
    """
    treated = max(0.0, float(treated_patients))
    arpp = max(0.0, float(arpp))
    years = max(0.0, float(treatment_years))
    disc = clamp(discount, 0.0, 1.0)
    funnel_cac = max(0.0, float(funnel_cac_total))
    platform = max(0.0, float(platform_costs))

    # Revenue calculations
    gross_revenue = treated * arpp * years
    net_revenue = gross_revenue * (1.0 - disc)
    
    # Cost calculations
    total_cost = funnel_cac + platform
    
    # Profit and ROI
    net_profit = net_revenue - total_cost
    roi = (net_revenue / total_cost) if total_cost > 0 else float("nan")
    
    # Cost per treated patient
    cost_per_patient = (total_cost / treated) if treated > 0 else float("nan")

    return {
        "treated_patients": treated,
        "gross_revenue": gross_revenue,
        "net_revenue": net_revenue,
        "discount": disc,
        "funnel_cac_total": funnel_cac,
        "platform_costs_total": platform,
        "total_cost": total_cost,
        "net_profit": net_profit,
        "roi": roi,
        "cost_per_patient": cost_per_patient,
    }


def run_scenario(scenario: dict, stage_names: List[str], base_population: Optional[float] = None):
    """
    Run a complete scenario calculation.
    
    Args:
        scenario: Scenario dictionary (baseline or dario)
        stage_names: List of stage names
        base_population: Override base population (if using shared)
    
    Returns:
        Tuple of (funnel_results, financials)
    """
    stages = build_stage_inputs(scenario, stage_names)
    pop = base_population if base_population is not None else float(scenario["base_population"])
    funnel_results = compute_funnel(stages, pop)
    platform_costs = sum(scenario.get("platform_costs", {}).values())
    
    financials = compute_financials(
        treated_patients=funnel_results[-1].patients,
        arpp=float(scenario["arpp"]),
        treatment_years=float(scenario["treatment_years"]),
        discount=float(scenario["discount"]),
        funnel_cac_total=funnel_results[-1].cumulative_cac,
        platform_costs=platform_costs,
    )
    
    return funnel_results, financials


def compute_incremental_metrics(baseline_fin: dict, dario_fin: dict) -> dict:
    """
    Calculate incremental/delta metrics between baseline and Dario scenarios.
    
    Formulas:
    - Incremental Treated Patients = Dario Patients - Baseline Patients
    - Incremental Net Revenue = Dario Net Revenue - Baseline Net Revenue
    - Incremental Cost = Dario Total Cost - Baseline Total Cost
    - Incremental Net Profit = Dario Net Profit - Baseline Net Profit
    - Incremental ROI (Revenue) = Incremental Net Revenue / Incremental Cost
    - Incremental ROI (Profit) = Incremental Net Profit / Incremental Cost
    - Cost per Incremental Patient = Incremental Cost / Incremental Patients
    
    Args:
        baseline_fin: Financial metrics for baseline scenario
        dario_fin: Financial metrics for Dario scenario
    
    Returns:
        Dictionary with all incremental metrics
    """
    # Incremental values
    incr_patients = dario_fin["treated_patients"] - baseline_fin["treated_patients"]
    incr_gross = dario_fin["gross_revenue"] - baseline_fin["gross_revenue"]
    incr_net_revenue = dario_fin["net_revenue"] - baseline_fin["net_revenue"]
    incr_cost = dario_fin["total_cost"] - baseline_fin["total_cost"]
    incr_profit = dario_fin["net_profit"] - baseline_fin["net_profit"]
    
    # ROI calculations (guard against divide-by-zero and negative costs)
    if incr_cost > 0:
        incr_roi_revenue = incr_net_revenue / incr_cost
        incr_roi_profit = incr_profit / incr_cost
    else:
        incr_roi_revenue = float("nan")
        incr_roi_profit = float("nan")
    
    # Cost per incremental patient
    if incr_patients > 0:
        cost_per_incr_patient = incr_cost / incr_patients
    else:
        cost_per_incr_patient = float("nan")
    
    return {
        "incremental_patients": incr_patients,
        "incremental_gross_revenue": incr_gross,
        "incremental_net_revenue": incr_net_revenue,
        "incremental_cost": incr_cost,
        "incremental_profit": incr_profit,
        "incremental_roi_revenue": incr_roi_revenue,
        "incremental_roi_profit": incr_roi_profit,
        "cost_per_incremental_patient": cost_per_incr_patient,
    }


def compute_breakeven(dario_fin: dict, incr: dict) -> dict:
    """
    Calculate break-even points for the Dario investment.
    
    Formulas:
    - Break-even Incremental Patients = Incremental Cost / Net Revenue per Patient (Dario)
    - Net Revenue per Patient = ARPP × Treatment Years × (1 - Discount)
    
    Args:
        dario_fin: Financial metrics for Dario scenario
        incr: Incremental metrics
    
    Returns:
        Dictionary with break-even calculations
    """
    # Net revenue per treated patient in Dario scenario
    if dario_fin["treated_patients"] > 0:
        net_rev_per_patient = dario_fin["net_revenue"] / dario_fin["treated_patients"]
    else:
        net_rev_per_patient = 0.0
    
    # Break-even incremental patients needed
    if net_rev_per_patient > 0:
        breakeven_patients = incr["incremental_cost"] / net_rev_per_patient
    else:
        breakeven_patients = float("nan")
    
    # Current patients above/below break-even
    patients_vs_breakeven = incr["incremental_patients"] - breakeven_patients if breakeven_patients == breakeven_patients else float("nan")
    
    return {
        "net_revenue_per_patient": net_rev_per_patient,
        "breakeven_incremental_patients": breakeven_patients,
        "patients_vs_breakeven": patients_vs_breakeven,
        "is_above_breakeven": patients_vs_breakeven > 0 if patients_vs_breakeven == patients_vs_breakeven else False,
    }


def compute_cac_sensitivity(
    model: dict,
    cac_min: float,
    cac_max: float,
    cac_step: float
) -> List[dict]:
    """
    Run sensitivity analysis varying baseline CAC (Stage 6).
    
    Args:
        model: Complete model dictionary
        cac_min: Minimum CAC to test
        cac_max: Maximum CAC to test
        cac_step: Step size between CAC values
    
    Returns:
        List of dictionaries with CAC value and resulting metrics
    """
    results = []
    stage_names = model["shared"]["stage_names"]
    
    # Get base population
    if model["shared"]["use_shared_base_population"]:
        base_pop = model["shared"]["shared_base_population"]
    else:
        base_pop = None
    
    # Dario scenario stays constant
    _, dario_fin = run_scenario(model["dario"], stage_names, base_pop)
    
    # Vary baseline CAC
    cac = cac_min
    while cac <= cac_max:
        # Create modified baseline scenario
        baseline_mod = copy.deepcopy(model["baseline"])
        baseline_mod["cac"][5] = cac  # Stage 6 CAC
        
        # Run baseline with modified CAC
        _, baseline_fin = run_scenario(baseline_mod, stage_names, base_pop)
        
        # Calculate incremental metrics
        incr = compute_incremental_metrics(baseline_fin, dario_fin)
        breakeven = compute_breakeven(dario_fin, incr)
        
        results.append({
            "baseline_cac": cac,
            "baseline_total_cost": baseline_fin["total_cost"],
            "baseline_net_profit": baseline_fin["net_profit"],
            "incremental_cost": incr["incremental_cost"],
            "incremental_profit": incr["incremental_profit"],
            "incremental_roi_revenue": incr["incremental_roi_revenue"],
            "incremental_roi_profit": incr["incremental_roi_profit"],
            "breakeven_patients": breakeven["breakeven_incremental_patients"],
        })
        
        cac += cac_step
    
    return results


def run_full_model(model: dict) -> dict:
    """
    Run complete model calculation including both scenarios and incremental metrics.
    
    Args:
        model: Complete model dictionary with shared, baseline, and dario
    
    Returns:
        Dictionary with all results
    """
    stage_names = model["shared"]["stage_names"]
    
    # Determine base population
    if model["shared"]["use_shared_base_population"]:
        base_pop = model["shared"]["shared_base_population"]
    else:
        base_pop = None
    
    # Run both scenarios
    baseline_funnel, baseline_fin = run_scenario(model["baseline"], stage_names, base_pop)
    dario_funnel, dario_fin = run_scenario(model["dario"], stage_names, base_pop)
    
    # Calculate incremental metrics
    incr = compute_incremental_metrics(baseline_fin, dario_fin)
    
    # Calculate break-even
    breakeven = compute_breakeven(dario_fin, incr)
    
    # Run sensitivity if baseline is in sensitivity mode
    sensitivity = None
    if model["baseline"].get("cac_mode") == "sensitivity":
        sens_cfg = model["baseline"]["cac_sensitivity"]
        sensitivity = compute_cac_sensitivity(
            model,
            sens_cfg["min"],
            sens_cfg["max"],
            sens_cfg["step"]
        )
    
    return {
        "baseline_funnel": baseline_funnel,
        "baseline_fin": baseline_fin,
        "dario_funnel": dario_funnel,
        "dario_fin": dario_fin,
        "incremental": incr,
        "breakeven": breakeven,
        "sensitivity": sensitivity,
    }


# =============================================================================
# EXCEL EXPORT
# =============================================================================
def build_excel_report(model_name: str, results: dict, model: dict) -> bytes:
    """
    Build a comprehensive Excel report with multiple sheets.
    
    Sheets:
    - Summary: Key metrics for both scenarios and incremental
    - Baseline Funnel: Stage-by-stage baseline data
    - Dario Funnel: Stage-by-stage Dario data
    - Incremental: Detailed incremental metrics
    - Sensitivity: CAC sensitivity analysis (if available)
    """
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    def style_header_row(ws, row=1):
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

    def set_col_widths(ws, widths):
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ----- Summary Sheet -----
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum["A1"] = f"PharmaROI Intelligence — {model_name}"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum.merge_cells("A1:D1")

    # Headers
    ws_sum["A3"] = "Metric"
    ws_sum["B3"] = "Baseline"
    ws_sum["C3"] = "Dario"
    ws_sum["D3"] = "Incremental"
    style_header_row(ws_sum, 3)

    baseline = results["baseline_fin"]
    dario = results["dario_fin"]
    incr = results["incremental"]

    summary_rows = [
        ("Treated Patients", baseline["treated_patients"], dario["treated_patients"], incr["incremental_patients"]),
        ("Gross Revenue", baseline["gross_revenue"], dario["gross_revenue"], incr["incremental_gross_revenue"]),
        ("Net Revenue", baseline["net_revenue"], dario["net_revenue"], incr["incremental_net_revenue"]),
        ("Total Cost", baseline["total_cost"], dario["total_cost"], incr["incremental_cost"]),
        ("Net Profit", baseline["net_profit"], dario["net_profit"], incr["incremental_profit"]),
        ("ROI", baseline["roi"], dario["roi"], incr["incremental_roi_profit"]),
    ]

    for i, (label, b_val, d_val, i_val) in enumerate(summary_rows):
        r = 4 + i
        ws_sum[f"A{r}"] = label
        ws_sum[f"B{r}"] = b_val if b_val == b_val else None
        ws_sum[f"C{r}"] = d_val if d_val == d_val else None
        ws_sum[f"D{r}"] = i_val if i_val == i_val else None

    set_col_widths(ws_sum, {1: 20, 2: 18, 3: 18, 4: 18})

    # ----- Baseline Funnel Sheet -----
    ws_base = wb.create_sheet("Baseline Funnel")
    headers = ["#", "Stage", "Ratio", "Patients", "CAC/pt", "Stage CAC", "Cumulative CAC"]
    for c, h in enumerate(headers, start=1):
        ws_base.cell(row=1, column=c, value=h)
    style_header_row(ws_base, 1)

    for r_idx, r in enumerate(results["baseline_funnel"], start=2):
        ws_base.cell(row=r_idx, column=1, value=r_idx-1)
        ws_base.cell(row=r_idx, column=2, value=r.name)
        ws_base.cell(row=r_idx, column=3, value=r.ratio_used)
        ws_base.cell(row=r_idx, column=4, value=r.patients)
        ws_base.cell(row=r_idx, column=5, value=r.cac_per_patient)
        ws_base.cell(row=r_idx, column=6, value=r.stage_cac)
        ws_base.cell(row=r_idx, column=7, value=r.cumulative_cac)

    set_col_widths(ws_base, {1: 5, 2: 45, 3: 10, 4: 15, 5: 12, 6: 15, 7: 18})

    # ----- Dario Funnel Sheet -----
    ws_dario = wb.create_sheet("Dario Funnel")
    for c, h in enumerate(headers, start=1):
        ws_dario.cell(row=1, column=c, value=h)
    style_header_row(ws_dario, 1)

    for r_idx, r in enumerate(results["dario_funnel"], start=2):
        ws_dario.cell(row=r_idx, column=1, value=r_idx-1)
        ws_dario.cell(row=r_idx, column=2, value=r.name)
        ws_dario.cell(row=r_idx, column=3, value=r.ratio_used)
        ws_dario.cell(row=r_idx, column=4, value=r.patients)
        ws_dario.cell(row=r_idx, column=5, value=r.cac_per_patient)
        ws_dario.cell(row=r_idx, column=6, value=r.stage_cac)
        ws_dario.cell(row=r_idx, column=7, value=r.cumulative_cac)

    set_col_widths(ws_dario, {1: 5, 2: 45, 3: 10, 4: 15, 5: 12, 6: 15, 7: 18})

    # ----- Incremental Sheet -----
    ws_incr = wb.create_sheet("Incremental")
    ws_incr["A1"] = "Incremental Metrics"
    ws_incr["A1"].font = Font(bold=True, size=12)

    incr_rows = [
        ("Incremental Treated Patients", incr["incremental_patients"]),
        ("Incremental Gross Revenue", incr["incremental_gross_revenue"]),
        ("Incremental Net Revenue", incr["incremental_net_revenue"]),
        ("Incremental Cost", incr["incremental_cost"]),
        ("Incremental Net Profit", incr["incremental_profit"]),
        ("Incremental ROI (Revenue)", incr["incremental_roi_revenue"]),
        ("Incremental ROI (Profit)", incr["incremental_roi_profit"]),
        ("Cost per Incremental Patient", incr["cost_per_incremental_patient"]),
        ("Break-even Incremental Patients", results["breakeven"]["breakeven_incremental_patients"]),
    ]

    ws_incr["A3"] = "Metric"
    ws_incr["B3"] = "Value"
    style_header_row(ws_incr, 3)

    for i, (label, val) in enumerate(incr_rows):
        r = 4 + i
        ws_incr[f"A{r}"] = label
        ws_incr[f"B{r}"] = val if val == val else None

    set_col_widths(ws_incr, {1: 35, 2: 20})

    # ----- Sensitivity Sheet (if available) -----
    if results["sensitivity"]:
        ws_sens = wb.create_sheet("Sensitivity")
        sens_headers = ["Baseline CAC", "Baseline Cost", "Incr Cost", "Incr Profit", "Incr ROI (Profit)"]
        for c, h in enumerate(sens_headers, start=1):
            ws_sens.cell(row=1, column=c, value=h)
        style_header_row(ws_sens, 1)

        for r_idx, row in enumerate(results["sensitivity"], start=2):
            ws_sens.cell(row=r_idx, column=1, value=row["baseline_cac"])
            ws_sens.cell(row=r_idx, column=2, value=row["baseline_total_cost"])
            ws_sens.cell(row=r_idx, column=3, value=row["incremental_cost"])
            ws_sens.cell(row=r_idx, column=4, value=row["incremental_profit"])
            ws_sens.cell(row=r_idx, column=5, value=row["incremental_roi_profit"])

        set_col_widths(ws_sens, {1: 15, 2: 18, 3: 15, 4: 15, 5: 18})

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# =============================================================================
# SESSION STATE INITIALIZATION
# =============================================================================
def init_session():
    """Initialize session state with default values and migrate old models."""
    if "models" not in st.session_state:
        st.session_state["models"] = [get_default_model()]
        st.session_state["model_names"] = ["Model 1"]
        st.session_state["active_model_idx"] = 0
    else:
        # Migrate any old-format models
        for i, model in enumerate(st.session_state["models"]):
            st.session_state["models"][i] = migrate_old_model(model)


init_session()


# =============================================================================
# PAGE CONFIGURATION
# =============================================================================
st.set_page_config(
    page_title="PharmaROI V4 — Baseline vs Dario",
    page_icon="📈",
    layout="wide"
)

st.title("PharmaROI Intelligence — V4")
st.caption("Compare Baseline (without Dario) vs Dario-Enabled scenarios and calculate incremental ROI.")


# =============================================================================
# MODEL MANAGEMENT BAR
# =============================================================================
mgmt_col1, mgmt_col2, mgmt_col3, mgmt_col4 = st.columns([2, 2, 2, 4])

with mgmt_col1:
    if st.button("Add New Model", use_container_width=True):
        n = len(st.session_state["models"]) + 1
        st.session_state["models"].append(get_default_model())
        st.session_state["model_names"].append(f"Model {n}")
        st.session_state["active_model_idx"] = len(st.session_state["models"]) - 1
        st.rerun()

with mgmt_col2:
    copy_options = st.session_state["model_names"]
    copy_source = st.selectbox(
        "Copy from:",
        options=range(len(copy_options)),
        format_func=lambda i: copy_options[i],
        index=st.session_state["active_model_idx"],
        key="copy_source_select",
        label_visibility="collapsed",
    )
    if st.button("Copy This Model", use_container_width=True):
        new_state = copy.deepcopy(st.session_state["models"][copy_source])
        new_name = st.session_state["model_names"][copy_source] + " (copy)"
        st.session_state["models"].append(new_state)
        st.session_state["model_names"].append(new_name)
        st.session_state["active_model_idx"] = len(st.session_state["models"]) - 1
        st.rerun()

with mgmt_col3:
    can_delete = len(st.session_state["models"]) > 1
    if "confirm_delete" not in st.session_state:
        st.session_state["confirm_delete"] = False

    if not st.session_state["confirm_delete"]:
        if st.button("Delete Current", use_container_width=True, disabled=not can_delete):
            st.session_state["confirm_delete"] = True
            st.rerun()
    else:
        idx = st.session_state["active_model_idx"]
        st.warning(f"Delete '{st.session_state['model_names'][idx]}'?")
        confirm_cols = st.columns(2)
        with confirm_cols[0]:
            if st.button("Yes, Delete", use_container_width=True, type="primary"):
                st.session_state["models"].pop(idx)
                st.session_state["model_names"].pop(idx)
                st.session_state["active_model_idx"] = max(0, idx - 1)
                st.session_state["confirm_delete"] = False
                st.rerun()
        with confirm_cols[1]:
            if st.button("Cancel", use_container_width=True):
                st.session_state["confirm_delete"] = False
                st.rerun()

with mgmt_col4:
    idx = st.session_state["active_model_idx"]
    new_name = st.text_input(
        "Rename current model:",
        value=st.session_state["model_names"][idx],
        key=f"rename_model_{idx}",
        label_visibility="collapsed",
        placeholder="Rename current model…",
    )
    if new_name != st.session_state["model_names"][idx]:
        st.session_state["model_names"][idx] = new_name


# =============================================================================
# TABS: ONE PER MODEL + COMPARISON
# =============================================================================
tab_labels = st.session_state["model_names"] + ["Comparison"]
tabs = st.tabs(tab_labels)


# =============================================================================
# MODEL TABS
# =============================================================================
for model_idx, model_tab in enumerate(tabs[:-1]):
    with model_tab:
        model = st.session_state["models"][model_idx]
        model_name = st.session_state["model_names"][model_idx]
        tab_color = TAB_PALETTE[model_idx % len(TAB_PALETTE)]

        # Update active model index when this tab is viewed
        st.session_state["active_model_idx"] = model_idx

        # =====================================================================
        # SHARED SETTINGS
        # =====================================================================
        with st.expander("Shared Model Settings", expanded=False):
            st.markdown("**Base Population**")
            model["shared"]["use_shared_base_population"] = st.checkbox(
                "Use same base population for both scenarios",
                value=model["shared"]["use_shared_base_population"],
                key=f"shared_pop_toggle_{model_idx}",
            )

            if model["shared"]["use_shared_base_population"]:
                model["shared"]["shared_base_population"] = st.number_input(
                    "Shared Base Population (TAM)",
                    min_value=0,
                    step=100_000,
                    value=int(model["shared"]["shared_base_population"]),
                    key=f"shared_pop_{model_idx}",
                )

            st.markdown("**Stage Names**")
            with st.expander("Customize Stage Names"):
                for sidx in range(NUM_STAGES):
                    model["shared"]["stage_names"][sidx] = st.text_input(
                        f"Stage {sidx+1}:",
                        value=model["shared"]["stage_names"][sidx],
                        key=f"stage_name_{model_idx}_{sidx}",
                    )

            # Reset buttons
            st.markdown("---")
            reset_col1, reset_col2 = st.columns(2)
            with reset_col1:
                if st.button("Reset to Defaults", key=f"reset_default_{model_idx}"):
                    st.session_state["models"][model_idx] = get_default_model()
                    st.rerun()
            with reset_col2:
                if st.button("Reset to Zero", key=f"reset_zero_{model_idx}"):
                    st.session_state["models"][model_idx] = get_zero_model()
                    st.rerun()

        # =====================================================================
        # BASELINE SCENARIO
        # =====================================================================
        st.markdown("---")
        st.subheader("Baseline Scenario (Without Dario)")
        st.caption("Configure the current state / status quo assumptions")

        baseline = model["baseline"]

        with st.expander("Baseline Settings", expanded=True):
            # Base population (if not shared)
            if not model["shared"]["use_shared_base_population"]:
                baseline["base_population"] = st.number_input(
                    "Baseline Base Population",
                    min_value=0,
                    step=100_000,
                    value=int(baseline["base_population"]),
                    key=f"baseline_pop_{model_idx}",
                )

            # Revenue settings
            st.markdown("**Revenue Assumptions**")
            b_col1, b_col2, b_col3 = st.columns(3)
            with b_col1:
                baseline["arpp"] = st.number_input(
                    "ARPP ($/year)",
                    min_value=0.0,
                    step=1000.0,
                    value=float(baseline["arpp"]),
                    key=f"baseline_arpp_{model_idx}",
                )
            with b_col2:
                baseline["treatment_years"] = st.slider(
                    "Treatment Years",
                    min_value=0.1,
                    max_value=5.0,
                    step=0.1,
                    value=float(baseline["treatment_years"]),
                    key=f"baseline_years_{model_idx}",
                )
            with b_col3:
                baseline["discount"] = st.slider(
                    "Discount (Gross→Net)",
                    min_value=0.0,
                    max_value=1.0,
                    step=0.01,
                    value=float(baseline["discount"]),
                    key=f"baseline_discount_{model_idx}",
                )

            # CAC Mode
            st.markdown("**CAC Configuration**")
            baseline["cac_mode"] = st.radio(
                "Baseline CAC Input Mode:",
                options=["direct", "sensitivity"],
                format_func=lambda x: "Direct Input" if x == "direct" else "Sensitivity Analysis",
                index=0 if baseline["cac_mode"] == "direct" else 1,
                key=f"baseline_cac_mode_{model_idx}",
                horizontal=True,
            )

            if baseline["cac_mode"] == "direct":
                # Direct CAC input for Stage 6
                baseline["cac"][5] = st.number_input(
                    "Stage 6 CAC ($ per patient)",
                    min_value=0.0,
                    step=1.0,
                    value=float(baseline["cac"][5]),
                    key=f"baseline_cac6_{model_idx}",
                    help="Customer acquisition cost at Stage 6 (Activation)",
                )
            else:
                # Sensitivity range
                st.markdown("**Sensitivity Range (Stage 6 CAC)**")
                sens_col1, sens_col2, sens_col3, sens_col4 = st.columns(4)
                with sens_col1:
                    baseline["cac_sensitivity"]["min"] = st.number_input(
                        "Min CAC",
                        min_value=0.0,
                        step=1.0,
                        value=float(baseline["cac_sensitivity"]["min"]),
                        key=f"baseline_sens_min_{model_idx}",
                    )
                with sens_col2:
                    baseline["cac_sensitivity"]["max"] = st.number_input(
                        "Max CAC",
                        min_value=0.0,
                        step=1.0,
                        value=float(baseline["cac_sensitivity"]["max"]),
                        key=f"baseline_sens_max_{model_idx}",
                    )
                with sens_col3:
                    baseline["cac_sensitivity"]["step"] = st.number_input(
                        "Step",
                        min_value=0.1,
                        step=0.5,
                        value=float(baseline["cac_sensitivity"]["step"]),
                        key=f"baseline_sens_step_{model_idx}",
                    )
                with sens_col4:
                    baseline["cac_sensitivity"]["base"] = st.number_input(
                        "Base Case",
                        min_value=0.0,
                        step=1.0,
                        value=float(baseline["cac_sensitivity"]["base"]),
                        key=f"baseline_sens_base_{model_idx}",
                    )
                # Use base case for direct calculations
                baseline["cac"][5] = baseline["cac_sensitivity"]["base"]

            # Funnel ratios
            st.markdown("**Funnel Conversion Rates**")
            with st.expander("Baseline Stage Ratios"):
                for sidx in range(1, NUM_STAGES):  # Skip stage 1 (always 1.0)
                    baseline["ratios"][sidx] = st.slider(
                        f"Stage {sidx+1}: {model['shared']['stage_names'][sidx][:40]}",
                        min_value=0.0,
                        max_value=1.0,
                        step=0.01,
                        value=float(baseline["ratios"][sidx]),
                        key=f"baseline_ratio_{model_idx}_{sidx}",
                    )

        # =====================================================================
        # DARIO SCENARIO
        # =====================================================================
        st.markdown("---")
        st.subheader("Dario-Enabled Scenario")
        st.caption("Configure the Dario-enabled assumptions (expected improvements)")

        dario = model["dario"]

        with st.expander("Dario Settings", expanded=True):
            # Base population (if not shared)
            if not model["shared"]["use_shared_base_population"]:
                dario["base_population"] = st.number_input(
                    "Dario Base Population",
                    min_value=0,
                    step=100_000,
                    value=int(dario["base_population"]),
                    key=f"dario_pop_{model_idx}",
                )

            # Revenue settings
            st.markdown("**Revenue Assumptions**")
            d_col1, d_col2, d_col3 = st.columns(3)
            with d_col1:
                dario["arpp"] = st.number_input(
                    "ARPP ($/year)",
                    min_value=0.0,
                    step=1000.0,
                    value=float(dario["arpp"]),
                    key=f"dario_arpp_{model_idx}",
                )
            with d_col2:
                dario["treatment_years"] = st.slider(
                    "Treatment Years",
                    min_value=0.1,
                    max_value=5.0,
                    step=0.1,
                    value=float(dario["treatment_years"]),
                    key=f"dario_years_{model_idx}",
                )
            with d_col3:
                dario["discount"] = st.slider(
                    "Discount (Gross→Net)",
                    min_value=0.0,
                    max_value=1.0,
                    step=0.01,
                    value=float(dario["discount"]),
                    key=f"dario_discount_{model_idx}",
                )

            # Dario CAC (Stage 6)
            st.markdown("**CAC Configuration**")
            dario["cac"][5] = st.number_input(
                "Stage 6 CAC ($ per patient)",
                min_value=0.0,
                step=1.0,
                value=float(dario["cac"][5]),
                key=f"dario_cac6_{model_idx}",
                help="Customer acquisition cost at Stage 6 with Dario",
            )

            # Platform costs
            st.markdown("**Platform Costs**")
            pc = dario["platform_costs"]
            pc_col1, pc_col2 = st.columns(2)
            with pc_col1:
                pc["dario_connect_config"] = st.number_input(
                    "Dario Connect Configuration",
                    min_value=0.0,
                    step=10000.0,
                    value=float(pc["dario_connect_config"]),
                    key=f"dario_pc1_{model_idx}",
                )
                pc["dario_care_config"] = st.number_input(
                    "Dario Care Configuration",
                    min_value=0.0,
                    step=10000.0,
                    value=float(pc["dario_care_config"]),
                    key=f"dario_pc2_{model_idx}",
                )
                pc["sub_dario_connect"] = st.number_input(
                    "Subscription — Dario Connect",
                    min_value=0.0,
                    step=10000.0,
                    value=float(pc["sub_dario_connect"]),
                    key=f"dario_pc3_{model_idx}",
                )
            with pc_col2:
                pc["sub_dario_care"] = st.number_input(
                    "Subscription — Dario Care",
                    min_value=0.0,
                    step=10000.0,
                    value=float(pc["sub_dario_care"]),
                    key=f"dario_pc4_{model_idx}",
                )
                pc["maintenance_support"] = st.number_input(
                    "Maintenance & Support",
                    min_value=0.0,
                    step=10000.0,
                    value=float(pc["maintenance_support"]),
                    key=f"dario_pc5_{model_idx}",
                )
            st.caption(f"**Total Platform Costs:** {money(sum(pc.values()))}")

            # Funnel ratios
            st.markdown("**Funnel Conversion Rates**")
            with st.expander("Dario Stage Ratios"):
                for sidx in range(1, NUM_STAGES):
                    dario["ratios"][sidx] = st.slider(
                        f"Stage {sidx+1}: {model['shared']['stage_names'][sidx][:40]}",
                        min_value=0.0,
                        max_value=1.0,
                        step=0.01,
                        value=float(dario["ratios"][sidx]),
                        key=f"dario_ratio_{model_idx}_{sidx}",
                    )

        # =====================================================================
        # RUN CALCULATIONS
        # =====================================================================
        results = run_full_model(model)
        baseline_fin = results["baseline_fin"]
        dario_fin = results["dario_fin"]
        incr = results["incremental"]
        breakeven = results["breakeven"]

        # =====================================================================
        # KPI SUMMARY CARDS
        # =====================================================================
        st.markdown("---")
        st.subheader("Results Summary")

        # Row 1: Treated Patients
        st.markdown("**Treated Patients**")
        kpi_col1, kpi_col2, kpi_col3 = st.columns(3)
        with kpi_col1:
            st.metric("Baseline", number(baseline_fin["treated_patients"]))
        with kpi_col2:
            st.metric("Dario", number(dario_fin["treated_patients"]))
        with kpi_col3:
            st.metric(
                "Incremental",
                number(incr["incremental_patients"]),
                delta=f"{incr['incremental_patients']:+,.0f}" if incr["incremental_patients"] != 0 else None,
            )

        # Row 2: Net Revenue
        st.markdown("**Net Revenue**")
        kpi_col4, kpi_col5, kpi_col6 = st.columns(3)
        with kpi_col4:
            st.metric("Baseline", money(baseline_fin["net_revenue"]))
        with kpi_col5:
            st.metric("Dario", money(dario_fin["net_revenue"]))
        with kpi_col6:
            st.metric(
                "Incremental",
                money(incr["incremental_net_revenue"]),
                delta=delta_fmt(incr["incremental_net_revenue"], "money") if incr["incremental_net_revenue"] != 0 else None,
            )

        # Row 3: Total Cost
        st.markdown("**Total Cost**")
        kpi_col7, kpi_col8, kpi_col9 = st.columns(3)
        with kpi_col7:
            st.metric("Baseline", money(baseline_fin["total_cost"]))
        with kpi_col8:
            st.metric("Dario", money(dario_fin["total_cost"]))
        with kpi_col9:
            st.metric(
                "Incremental",
                money(incr["incremental_cost"]),
                delta=delta_fmt(incr["incremental_cost"], "money") if incr["incremental_cost"] != 0 else None,
            )

        # Row 4: Net Profit & ROI
        st.markdown("**Profitability**")
        kpi_col10, kpi_col11, kpi_col12, kpi_col13 = st.columns(4)
        with kpi_col10:
            st.metric("Baseline Net Profit", money(baseline_fin["net_profit"]))
        with kpi_col11:
            st.metric("Dario Net Profit", money(dario_fin["net_profit"]))
        with kpi_col12:
            st.metric("Incremental Profit", money(incr["incremental_profit"]))
        with kpi_col13:
            st.metric("Incremental ROI", roix(incr["incremental_roi_profit"]))

        # Row 5: Break-even
        st.markdown("**Break-even Analysis**")
        be_col1, be_col2, be_col3 = st.columns(3)
        with be_col1:
            st.metric("Break-even Patients Needed", number(breakeven["breakeven_incremental_patients"]))
        with be_col2:
            st.metric("Actual Incremental Patients", number(incr["incremental_patients"]))
        with be_col3:
            if breakeven["is_above_breakeven"]:
                st.success(f"Above break-even by {number(breakeven['patients_vs_breakeven'])} patients")
            else:
                st.warning(f"Below break-even by {number(abs(breakeven['patients_vs_breakeven']))} patients")

        # =====================================================================
        # FUNNEL TABLES
        # =====================================================================
        st.markdown("---")
        st.subheader("Funnel Comparison")

        if pd is not None:
            funnel_tab1, funnel_tab2 = st.tabs(["Baseline Funnel", "Dario Funnel"])

            with funnel_tab1:
                baseline_rows = []
                for ridx, r in enumerate(results["baseline_funnel"]):
                    baseline_rows.append({
                        "#": ridx + 1,
                        "Stage": r.name,
                        "Ratio": pct(r.ratio_used) if ridx > 0 else "—",
                        "Patients": number(r.patients),
                        "CAC/pt": money(r.cac_per_patient),
                        "Cumulative CAC": money(r.cumulative_cac),
                    })
                st.dataframe(pd.DataFrame(baseline_rows), use_container_width=True, hide_index=True)

            with funnel_tab2:
                dario_rows = []
                for ridx, r in enumerate(results["dario_funnel"]):
                    dario_rows.append({
                        "#": ridx + 1,
                        "Stage": r.name,
                        "Ratio": pct(r.ratio_used) if ridx > 0 else "—",
                        "Patients": number(r.patients),
                        "CAC/pt": money(r.cac_per_patient),
                        "Cumulative CAC": money(r.cumulative_cac),
                    })
                st.dataframe(pd.DataFrame(dario_rows), use_container_width=True, hide_index=True)

        # =====================================================================
        # CHARTS
        # =====================================================================
        st.markdown("---")
        st.subheader("Visual Comparisons")

        if pd is not None:
            # Side-by-side bar chart
            st.markdown("**Baseline vs Dario: Key Metrics**")
            comparison_data = pd.DataFrame([
                {"Metric": "Treated Patients", "Scenario": "Baseline", "Value": baseline_fin["treated_patients"]},
                {"Metric": "Treated Patients", "Scenario": "Dario", "Value": dario_fin["treated_patients"]},
                {"Metric": "Net Revenue", "Scenario": "Baseline", "Value": baseline_fin["net_revenue"]},
                {"Metric": "Net Revenue", "Scenario": "Dario", "Value": dario_fin["net_revenue"]},
                {"Metric": "Total Cost", "Scenario": "Baseline", "Value": baseline_fin["total_cost"]},
                {"Metric": "Total Cost", "Scenario": "Dario", "Value": dario_fin["total_cost"]},
                {"Metric": "Net Profit", "Scenario": "Baseline", "Value": baseline_fin["net_profit"]},
                {"Metric": "Net Profit", "Scenario": "Dario", "Value": dario_fin["net_profit"]},
            ])

            scenario_colors = alt.Scale(
                domain=["Baseline", "Dario"],
                range=[COLORS["baseline"], COLORS["dario"]]
            )

            bar_chart = alt.Chart(comparison_data).mark_bar().encode(
                x=alt.X("Scenario:N", title=None),
                y=alt.Y("Value:Q", title="Value"),
                color=alt.Color("Scenario:N", scale=scenario_colors, legend=None),
                column=alt.Column("Metric:N", title=None),
                tooltip=["Scenario", alt.Tooltip("Value:Q", format=",.0f")],
            ).properties(width=150, height=300)

            st.altair_chart(bar_chart)

            # Stage-by-stage patient comparison
            st.markdown("**Funnel Stage Comparison: Patients by Stage**")
            stage_comparison = []
            for ridx, (b_res, d_res) in enumerate(zip(results["baseline_funnel"], results["dario_funnel"])):
                stage_comparison.append({
                    "Stage": f"{ridx+1}. {b_res.name[:30]}",
                    "Scenario": "Baseline",
                    "Patients": b_res.patients,
                })
                stage_comparison.append({
                    "Stage": f"{ridx+1}. {d_res.name[:30]}",
                    "Scenario": "Dario",
                    "Patients": d_res.patients,
                })

            stage_df = pd.DataFrame(stage_comparison)
            stage_chart = alt.Chart(stage_df).mark_line(point=True).encode(
                x=alt.X("Stage:N", sort=None, title=None, axis=alt.Axis(labelAngle=-45, labelLimit=200)),
                y=alt.Y("Patients:Q", title="Patients", axis=alt.Axis(format=",.0f")),
                color=alt.Color("Scenario:N", scale=scenario_colors),
                tooltip=["Stage", "Scenario", alt.Tooltip("Patients:Q", format=",.0f")],
            ).properties(height=400)

            st.altair_chart(stage_chart, use_container_width=True)

            # Sensitivity chart (if available)
            if results["sensitivity"]:
                st.markdown("**Sensitivity Analysis: Incremental ROI vs Baseline CAC**")
                sens_df = pd.DataFrame(results["sensitivity"])

                sens_chart = alt.Chart(sens_df).mark_line(point=True, color=COLORS["incremental"]).encode(
                    x=alt.X("baseline_cac:Q", title="Baseline CAC ($/patient)"),
                    y=alt.Y("incremental_roi_profit:Q", title="Incremental ROI (Profit)"),
                    tooltip=[
                        alt.Tooltip("baseline_cac:Q", title="Baseline CAC", format="$.0f"),
                        alt.Tooltip("incremental_roi_profit:Q", title="Incr ROI", format=".2f"),
                        alt.Tooltip("incremental_profit:Q", title="Incr Profit", format="$,.0f"),
                    ],
                ).properties(height=350)

                # Add break-even line at ROI = 1.0
                rule = alt.Chart(pd.DataFrame({"y": [1.0]})).mark_rule(
                    strokeDash=[5, 5],
                    color=COLORS["warning"]
                ).encode(y="y:Q")

                st.altair_chart(sens_chart + rule, use_container_width=True)

                # Sensitivity table
                with st.expander("Sensitivity Data Table"):
                    sens_display = sens_df.copy()
                    sens_display["baseline_cac"] = sens_display["baseline_cac"].map(lambda x: f"${x:,.0f}")
                    sens_display["incremental_cost"] = sens_display["incremental_cost"].map(lambda x: f"${x:,.0f}")
                    sens_display["incremental_profit"] = sens_display["incremental_profit"].map(lambda x: f"${x:,.0f}")
                    sens_display["incremental_roi_profit"] = sens_display["incremental_roi_profit"].map(lambda x: f"{x:.2f}x")
                    st.dataframe(sens_display[["baseline_cac", "incremental_cost", "incremental_profit", "incremental_roi_profit"]], use_container_width=True, hide_index=True)

        # =====================================================================
        # EXPORT
        # =====================================================================
        st.markdown("---")
        st.subheader("Export")

        export_col1, export_col2 = st.columns(2)
        with export_col1:
            xlsx_bytes = build_excel_report(model_name, results, model)
            st.download_button(
                "Download Excel Report",
                data=xlsx_bytes,
                file_name=f"{model_name.replace(' ', '_')}_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_xlsx_{model_idx}",
            )

        with export_col2:
            # Export model config as JSON
            model_export = {
                "name": model_name,
                "model": model,
            }
            json_str = json.dumps(model_export, indent=2)
            st.download_button(
                "Export Model Config (JSON)",
                data=json_str,
                file_name=f"{model_name.replace(' ', '_')}_config.json",
                mime="application/json",
                key=f"dl_json_{model_idx}",
            )


# =============================================================================
# COMPARISON TAB
# =============================================================================
with tabs[-1]:
    st.subheader("Model Comparison")

    if len(st.session_state["models"]) < 2:
        st.info("Add at least 2 models to compare them here.")
    else:
        # Model selection
        st.markdown("**Select models to compare:**")
        selected_model_names = st.multiselect(
            "Choose models",
            options=st.session_state["model_names"],
            default=st.session_state["model_names"],
            key="comparison_model_select",
            label_visibility="collapsed",
        )

        if len(selected_model_names) < 2:
            st.warning("Please select at least 2 models to compare.")
            st.stop()

        # Filter selected models
        selected_indices = [i for i, name in enumerate(st.session_state["model_names"]) if name in selected_model_names]
        selected_models = [st.session_state["models"][i] for i in selected_indices]
        selected_names = [st.session_state["model_names"][i] for i in selected_indices]

        # Build comparison data
        comparison_rows = []
        for mstate, mname in zip(selected_models, selected_names):
            results = run_full_model(mstate)
            baseline_fin = results["baseline_fin"]
            dario_fin = results["dario_fin"]
            incr = results["incremental"]

            comparison_rows.append({
                "Model": mname,
                "Baseline Patients": baseline_fin["treated_patients"],
                "Dario Patients": dario_fin["treated_patients"],
                "Incremental Patients": incr["incremental_patients"],
                "Baseline Net Revenue": baseline_fin["net_revenue"],
                "Dario Net Revenue": dario_fin["net_revenue"],
                "Incremental Net Revenue": incr["incremental_net_revenue"],
                "Baseline Cost": baseline_fin["total_cost"],
                "Dario Cost": dario_fin["total_cost"],
                "Incremental Cost": incr["incremental_cost"],
                "Incremental Profit": incr["incremental_profit"],
                "Incremental ROI": incr["incremental_roi_profit"],
            })

        if pd is not None:
            comp_df = pd.DataFrame(comparison_rows)

            # Summary table
            st.markdown("### Key Metrics")
            disp = comp_df.copy()
            for col in ["Baseline Patients", "Dario Patients", "Incremental Patients"]:
                disp[col] = disp[col].map(lambda x: f"{x:,.0f}")
            for col in ["Baseline Net Revenue", "Dario Net Revenue", "Incremental Net Revenue",
                        "Baseline Cost", "Dario Cost", "Incremental Cost", "Incremental Profit"]:
                disp[col] = disp[col].map(lambda x: f"${x:,.0f}")
            disp["Incremental ROI"] = disp["Incremental ROI"].map(lambda x: f"{x:.2f}x" if x == x else "—")

            st.dataframe(disp, use_container_width=True, hide_index=True)

            # Charts
            st.markdown("### Comparison Charts")

            model_color_scale = alt.Scale(
                domain=selected_names,
                range=TAB_PALETTE[:len(selected_names)],
            )

            chart_col1, chart_col2 = st.columns(2)

            with chart_col1:
                st.markdown("**Incremental ROI by Model**")
                roi_chart = alt.Chart(comp_df).mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4).encode(
                    x=alt.X("Model:N", title=None, axis=alt.Axis(labelAngle=-20)),
                    y=alt.Y("Incremental ROI:Q", title="ROI (x)"),
                    color=alt.Color("Model:N", scale=model_color_scale, legend=None),
                    tooltip=["Model:N", alt.Tooltip("Incremental ROI:Q", format=".2f")],
                ).properties(height=300)
                st.altair_chart(roi_chart, use_container_width=True)

            with chart_col2:
                st.markdown("**Incremental Profit by Model**")
                profit_chart = alt.Chart(comp_df).mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4).encode(
                    x=alt.X("Model:N", title=None, axis=alt.Axis(labelAngle=-20)),
                    y=alt.Y("Incremental Profit:Q", title="USD", axis=alt.Axis(format="$,.0f")),
                    color=alt.Color("Model:N", scale=model_color_scale, legend=None),
                    tooltip=["Model:N", alt.Tooltip("Incremental Profit:Q", format="$,.0f")],
                ).properties(height=300)
                st.altair_chart(profit_chart, use_container_width=True)

            chart_col3, chart_col4 = st.columns(2)

            with chart_col3:
                st.markdown("**Incremental Patients by Model**")
                patients_chart = alt.Chart(comp_df).mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4).encode(
                    x=alt.X("Model:N", title=None, axis=alt.Axis(labelAngle=-20)),
                    y=alt.Y("Incremental Patients:Q", title="Patients", axis=alt.Axis(format=",.0f")),
                    color=alt.Color("Model:N", scale=model_color_scale, legend=None),
                    tooltip=["Model:N", alt.Tooltip("Incremental Patients:Q", format=",.0f")],
                ).properties(height=300)
                st.altair_chart(patients_chart, use_container_width=True)

            with chart_col4:
                st.markdown("**Incremental Net Revenue by Model**")
                rev_chart = alt.Chart(comp_df).mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4).encode(
                    x=alt.X("Model:N", title=None, axis=alt.Axis(labelAngle=-20)),
                    y=alt.Y("Incremental Net Revenue:Q", title="USD", axis=alt.Axis(format="$,.0f")),
                    color=alt.Color("Model:N", scale=model_color_scale, legend=None),
                    tooltip=["Model:N", alt.Tooltip("Incremental Net Revenue:Q", format="$,.0f")],
                ).properties(height=300)
                st.altair_chart(rev_chart, use_container_width=True)

            # Export comparison
            st.markdown("### Export")
            comp_csv = comp_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                "Download Comparison CSV",
                data=comp_csv,
                file_name="pharmaroi_comparison.csv",
                mime="text/csv",
            )


# =============================================================================
# FOOTER / HELP
# =============================================================================
st.divider()
st.subheader("How to Interpret")
st.write("""
**Scenarios:**
- **Baseline**: Current state without Dario intervention
- **Dario**: State with Dario platform enabled (expected improvements)
- **Incremental**: The difference/lift that Dario provides

**Key Formulas:**
- **Gross Revenue** = Treated Patients × ARPP × Treatment Years
- **Net Revenue** = Gross Revenue × (1 - Discount)
- **Total Cost** = Funnel CAC + Platform Costs
- **Net Profit** = Net Revenue - Total Cost
- **Incremental ROI** = Incremental Net Profit / Incremental Cost

**Break-even:**
- The number of incremental patients needed for Dario investment to pay off
- If actual incremental patients > break-even patients, the investment is profitable

**Sensitivity Analysis:**
- When baseline CAC is uncertain, sensitivity mode shows how ROI changes across a range of assumptions
- The dotted line at ROI = 1.0x represents break-even
""")

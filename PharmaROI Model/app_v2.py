# app.py
# PharmaROI Intelligence — V3 (Multi-Model Comparison)
# Run: streamlit run "PharmaROI Model/app_v2.py"

from __future__ import annotations

import copy
import io
from dataclasses import dataclass
from typing import List

import streamlit as st

try:
    import pandas as pd
except Exception:
    pd = None

import plotly.express as px
import plotly.graph_objects as go

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# -----------------------------
# Color palette
# -----------------------------
COLORS = {
    "primary": "#0F6CBD",
    "revenue": "#0F6CBD",
    "costs": "#9CA3AF",
    "profit": "#10B981",
    "warning": "#F59E0B",
    "danger": "#EF4444",
    "muted": "#6B7280",
}

TAB_PALETTE = [
    "#0F6CBD", "#10B981", "#F59E0B", "#EF4444",
    "#8B5CF6", "#EC4899", "#06B6D4", "#84CC16",
]

# -----------------------------
# Funnel definitions
# -----------------------------
STAGE_NAMES: List[str] = [
    "Total Addressable Market for MASH",
    "F2 and F3",
    "MASH patients diagnosed",
    "Madrigal access to MASH patients",
    "Frequent users of online and social media resources",
    "Activation within 90 days mo onto Dario Connect for MASH",
    "Schedule telemedicine appointment",
    "Keep telemedicine appointment",
    "Obtain prescription for biopsy",
    "Get biopsy lab test",
    "Get positive lab results",
    "Complete post lab result consultation",
    "Get prescription for Rezdiffra",
]

SPONSOR_DEFAULTS = {
    "base_population": 10_000_000,
    "ratios": [1.00, 0.35, 0.16, 0.22, 0.75, 0.40, 0.15, 0.80, 1.00, 0.75, 0.90, 0.50, 0.90],
    "cac": [0.0, 0.0, 0.0, 0.0, 0.0, 10.0, 67.0, 83.0, 83.0, 111.0, 123.0, 247.0, 274.0],
    "arpp": 47_400.0,
    "treatment_years": 1.0,
    "discount": 0.68,
    "stage_active": [True] * len(STAGE_NAMES),
    "stage_names": STAGE_NAMES[:],
    "platform_costs": {
        "dario_connect_config": 500_000.0,
        "dario_care_config": 500_000.0,
        "sub_dario_connect": 1_000_000.0,
        "sub_dario_care": 250_000.0,
        "maintenance_support": 250_000.0,
    },
}

ZERO_SAMPLE = {
    "base_population": 0,
    "ratios": [0.0] * len(STAGE_NAMES),
    "cac": [0.0] * len(STAGE_NAMES),
    "arpp": 0.0,
    "treatment_years": 1.0,
    "discount": 0.0,
    "stage_active": [True] * len(STAGE_NAMES),
    "stage_names": ["Insert Stage Name"] * len(STAGE_NAMES),
    "platform_costs": {
        "dario_connect_config": 0.0,
        "dario_care_config": 0.0,
        "sub_dario_connect": 0.0,
        "sub_dario_care": 0.0,
        "maintenance_support": 0.0,
    },
}

# -----------------------------
# Formatting helpers
# -----------------------------
def clamp(x, lo, hi):
    return max(lo, min(hi, float(x)))

def money(x):
    return f"${x:,.0f}"

def number(x):
    return f"{x:,.0f}"

def pct(x):
    return f"{x * 100:,.1f}%"

def roix(x):
    return f"{x:,.2f}x"

# -----------------------------
# Core computations
# -----------------------------
@dataclass(frozen=True)
class StageInput:
    name: str
    active: bool
    ratio: float
    cac: float

@dataclass(frozen=True)
class StageResult:
    name: str
    active: bool
    ratio_used: float
    patients: float
    cac_per_patient: float
    stage_cac: float
    cumulative_cac: float


def compute_funnel(stages: List[StageInput], base_population: float) -> List[StageResult]:
    results = []
    prev_patients = max(0.0, float(base_population))
    total_cac_pool = 0.0

    for idx, s in enumerate(stages):
        if idx == 0:
            patients = prev_patients
            ratio_used = 1.0
        else:
            ratio_used = 1.0 if not s.active else clamp(s.ratio, 0.0, 1.0)
            patients = prev_patients * ratio_used

        if idx < 5:
            cac_pp = 0.0
            stage_cac = 0.0
            cumulative = 0.0
        elif idx == 5:
            cac_pp = 0.0 if not s.active else max(0.0, float(s.cac))
            stage_cac = patients * cac_pp
            total_cac_pool = stage_cac
            cumulative = total_cac_pool
        else:
            cumulative = total_cac_pool
            cac_pp = (total_cac_pool / patients) if patients > 0 else 0.0
            stage_cac = cac_pp * patients

        results.append(StageResult(
            name=s.name,
            active=s.active,
            ratio_used=ratio_used,
            patients=patients,
            cac_per_patient=cac_pp,
            stage_cac=stage_cac,
            cumulative_cac=cumulative,
        ))
        prev_patients = patients

    return results


def compute_financials(
    treated_patients,
    arpp,
    treatment_years,
    discount,
    funnel_cac_total,
    platform_costs=0.0,
):
    treated = max(0.0, float(treated_patients))
    arpp = max(0.0, float(arpp))
    years = max(0.0, float(treatment_years))
    disc = clamp(discount, 0.0, 1.0)
    funnel_cac = max(0.0, float(funnel_cac_total))
    platform = max(0.0, float(platform_costs))

    gross = treated * arpp * years
    net = gross * (1.0 - disc)
    net_profit = net - funnel_cac - platform
    roi = (net / (funnel_cac + platform)) if (funnel_cac + platform) > 0 else float("nan")

    return {
        "treated_patients": treated,
        "gross_revenue": gross,
        "net_revenue": net,
        "discount": disc,
        "funnel_cac_total": funnel_cac,
        "platform_costs_total": platform,
        "net_profit": net_profit,
        "roi_net": roi,
    }


def run_model(state: dict):
    stage_names = state.get("stage_names", STAGE_NAMES)
    stages = []
    for idx, name in enumerate(stage_names):
        stages.append(StageInput(
            name=name,
            active=bool(state["stage_active"][idx]),
            ratio=float(state["ratios"][idx]) if idx > 0 else 1.0,
            cac=float(state["cac"][idx]),
        ))
    base_pop = float(state["base_population"])
    funnel_results = compute_funnel(stages, base_pop)
    platform_costs = sum(state.get("platform_costs", {}).values())
    fin = compute_financials(
        treated_patients=funnel_results[-1].patients,
        arpp=float(state["arpp"]),
        treatment_years=float(state["treatment_years"]),
        discount=float(state["discount"]),
        funnel_cac_total=funnel_results[-1].cumulative_cac,
        platform_costs=platform_costs,
    )
    return funnel_results, fin

# -----------------------------
# Optimization phase helper
# -----------------------------
def build_phase_optimization_df(fin: dict, state: dict,
                                 eff_0_3: float = 1.0,
                                 eff_3_6: float = 1.0,
                                 eff_6_plus: float = 1.0):
    if pd is None:
        return None

    phased_enabled = bool(state.get("phased_enabled", False))

    if not phased_enabled:
        eff_0_3 = 1.0
        eff_3_6 = 1.0
        eff_6_plus = 1.0

    base_roi = float(fin["roi_net"]) if fin["roi_net"] == fin["roi_net"] else 0.0
    base_net_revenue = float(fin["net_revenue"])

    rows = [
        {
            "Phase": "Months 0-3",
            "Phase Order": 1,
            "Efficiency": eff_0_3,
            "ROI": base_roi * eff_0_3,
            "Net Revenue": base_net_revenue * eff_0_3,
        },
        {
            "Phase": "Months 3-6",
            "Phase Order": 2,
            "Efficiency": eff_3_6,
            "ROI": base_roi * eff_3_6,
            "Net Revenue": base_net_revenue * eff_3_6,
        },
        {
            "Phase": "Months 6-12",
            "Phase Order": 3,
            "Efficiency": eff_6_plus,
            "ROI": base_roi * eff_6_plus,
            "Net Revenue": base_net_revenue * eff_6_plus,
        },
    ]

    return pd.DataFrame(rows)

# -----------------------------
# Plotly chart helpers
# -----------------------------
def plotly_waterfall(fin):
    gross = fin["gross_revenue"]
    discount_amount = fin["gross_revenue"] - fin["net_revenue"]
    net_revenue = fin["net_revenue"]
    funnel_cac = fin["funnel_cac_total"]
    platform_costs = fin["platform_costs_total"]
    net_profit = fin["net_profit"]

    fig = go.Figure(go.Waterfall(
        name="Financial Bridge",
        orientation="v",
        measure=["relative", "relative", "total", "relative", "relative", "total"],
        x=[
            "Gross Revenue Potential",
            "Gross-to-Net Discount",
            "Net Revenue",
            "Patient Acquisition Cost",
            "Platform Investment",
            "Net Profit",
        ],
        text=[
            money(gross),
            f"-{money(discount_amount)}",
            money(net_revenue),
            f"-{money(funnel_cac)}",
            f"-{money(platform_costs)}",
            money(net_profit),
        ],
        textposition="outside",
        y=[gross, -discount_amount, 0, -funnel_cac, -platform_costs, 0],
        connector={"line": {"color": COLORS["muted"]}},
        increasing={"marker": {"color": COLORS["revenue"]}},
        decreasing={"marker": {"color": COLORS["danger"]}},
        totals={"marker": {"color": COLORS["profit"]}},
    ))
    fig.update_layout(
        title="Revenue-to-Profit Bridge",
        height=420,
        margin=dict(l=10, r=10, t=55, b=10),
        showlegend=False,
        yaxis_title="USD",
    )
    return fig


def plotly_funnel_patients(df_funnel, tab_color):
    plot_df = df_funnel.copy()
    plot_df["Patients Plot"] = plot_df["Patients"].apply(lambda x: max(float(x), 1.0))
    plot_df["Patients Label"] = plot_df["Patients"].apply(lambda x: f"{float(x):,.0f}")

    fig = px.bar(
        plot_df,
        x="Patients Plot",
        y="Stage",
        orientation="h",
        text="Patients Label",
    )
    fig.update_traces(
        marker_color=tab_color,
        textposition="outside",
        cliponaxis=False,
        hovertemplate="Stage: %{y}<br>Patients: %{text}<extra></extra>",
    )
    fig.update_layout(
        title="Patient Volume by Funnel Stage",
        height=500,
        margin=dict(l=10, r=40, t=55, b=10),
        xaxis_title="Patients (log scale)",
        yaxis_title=None,
        showlegend=False,
    )
    fig.update_xaxes(type="log")
    fig.update_yaxes(categoryorder="array", categoryarray=list(plot_df["Stage"])[::-1])
    return fig


def plotly_comparison_bar(comp_df, y_col, title, y_title, color_map):
    fig = px.bar(comp_df, x="Model", y=y_col, color="Model", color_discrete_map=color_map, text=y_col)

    if "ROI" in y_col:
        fig.update_traces(texttemplate="%{text:.2f}x", textposition="outside")
    elif "Discount" in y_col:
        fig.update_traces(texttemplate="%{text:.1%}", textposition="outside")
    elif "Patients" in y_col:
        fig.update_traces(texttemplate="%{text:,.0f}", textposition="outside")
    else:
        fig.update_traces(texttemplate="$%{text:,.0f}", textposition="outside")

    fig.update_layout(
        title=title,
        height=340,
        margin=dict(l=10, r=10, t=55, b=10),
        xaxis_title=None,
        yaxis_title=y_title,
        showlegend=False,
    )
    return fig

def plotly_phase_step_chart(df_phase, y_col, title, y_title, line_color):
    fig = go.Figure()

    phase_ranges = {
        "Months 0-3":  (0, 3),
        "Months 3-6":  (3, 6),
        "Months 6-12": (6, 12),
    }

    x_vals = []
    y_vals = []

    for _, row in df_phase.iterrows():
        phase = row["Phase"]
        x_start, x_end = phase_ranges.get(phase, (0, 12))
        y_val = row[y_col]
        x_vals.extend([x_start, x_end])
        y_vals.extend([y_val, y_val])

    fig.add_trace(go.Scatter(
        x=x_vals,
        y=y_vals,
        mode="lines+markers",
        line=dict(color=line_color, width=3),
        marker=dict(size=8, color=line_color),
        hovertemplate=f"Month: %{{x}}<br>{y_title}: %{{y:,.2f}}x<extra></extra>" if y_col == "ROI"
        else f"Month: %{{x}}<br>{y_title}: $%{{y:,.0f}}<extra></extra>",
        showlegend=False,
    ))

    fig.update_layout(
        title=title,
        height=360,
        margin=dict(l=10, r=10, t=55, b=10),
        xaxis=dict(title="Month", tickvals=list(range(0, 13)), range=[0, 12]),
        yaxis_title=y_title,
        showlegend=False,
        hovermode="x unified",
    )
    return fig

def plotly_phase_comparison_chart(df_phase_comp, y_col, title, y_title, color_map):
    fig = go.Figure()

    phase_ranges = {
        "Months 0-3":  (0, 3),
        "Months 3-6":  (3, 6),
        "Months 6-12": (6, 12),
    }

    for model_name in df_phase_comp["Model"].unique():
        model_df = df_phase_comp[df_phase_comp["Model"] == model_name]
        color = color_map.get(model_name, "#0F6CBD")

        x_vals = []
        y_vals = []

        for _, row in model_df.iterrows():
            phase = row["Phase"]
            x_start, x_end = phase_ranges.get(phase, (0, 12))
            y_val = row[y_col]
            x_vals.extend([x_start, x_end])
            y_vals.extend([y_val, y_val])

        fig.add_trace(go.Scatter(
            x=x_vals,
            y=y_vals,
            mode="lines+markers",
            name=model_name,
            line=dict(color=color, width=3),
            marker=dict(size=8, color=color),
            hovertemplate=f"{model_name}<br>Month: %{{x}}<br>{y_title}: %{{y:.2f}}x<extra></extra>" if y_col == "ROI"
            else f"{model_name}<br>Month: %{{x}}<br>{y_title}: $%{{y:,.0f}}<extra></extra>",
        ))

    fig.update_layout(
        title=title,
        height=380,
        margin=dict(l=10, r=10, t=55, b=10),
        xaxis=dict(title="Month", tickvals=list(range(0, 13)), range=[0, 12]),
        yaxis_title=y_title,
        legend_title=None,
        hovermode="x unified",
    )
    return fig

def plotly_per_patient_costs(df_pp_costs, color_map):
    fig = px.bar(
        df_pp_costs,
        x="Metric",
        y="Cost per Treated Patient",
        color="Model",
        barmode="group",
        color_discrete_map=color_map,
        text="Cost per Treated Patient",
        category_orders={
            "Metric": [
                "Funnel CAC per Treated Patient",
                "Platform Costs per Treated Patient",
                "Total Cost per Treated Patient",
            ]
        },
    )
    fig.update_traces(texttemplate="$%{text:,.0f}", textposition="outside")
    fig.update_layout(
        title="Per-Patient Cost Comparison by Scenario",
        height=380,
        margin=dict(l=10, r=10, t=55, b=10),
        xaxis_title=None,
        yaxis_title="Cost per Treated Patient",
        legend_title=None,
    )
    return fig

# -----------------------------
# Excel export helpers
# -----------------------------
def build_polished_excel_report(df_funnel, fin, colors):
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    muted_font = Font(color="6B7280")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    def set_col_widths(ws, widths):
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    def style_header_row(ws, row=1):
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum["A1"] = "PharmaROI Intelligence — Sponsor Summary"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum.merge_cells("A1:D1")

    summary_rows = [
        ("Treated Patients", fin["treated_patients"], "0"),
        ("Gross Revenue", fin["gross_revenue"], "$#,##0"),
        ("Discount", fin["discount"], "0.0%"),
        ("Net Revenue", fin["net_revenue"], "$#,##0"),
        ("Funnel CAC Total", fin["funnel_cac_total"], "$#,##0"),
        ("Platform Costs", fin["platform_costs_total"], "$#,##0"),
        ("Net Profit", fin["net_profit"], "$#,##0"),
        ("ROI (Net)", fin["roi_net"], "0.00x"),
    ]

    ws_sum["A3"] = "Metric"
    ws_sum["B3"] = "Value"
    ws_sum["C3"] = "Format"
    ws_sum["D3"] = "Notes"
    style_header_row(ws_sum, 3)

    start_row = 4
    for i, (label, value, fmt) in enumerate(summary_rows):
        r = start_row + i
        ws_sum[f"A{r}"] = label
        ws_sum[f"B{r}"] = float(value) if value == value else None
        ws_sum[f"C{r}"] = fmt
        ws_sum[f"D{r}"] = ""
        ws_sum[f"A{r}"].font = bold_font if label in ("Net Revenue", "ROI (Net)", "Net Profit") else Font()
        ws_sum[f"A{r}"].alignment = left
        ws_sum[f"B{r}"].alignment = left
        ws_sum[f"C{r}"].font = muted_font
        ws_sum[f"B{r}"].number_format = fmt

    ws_sum.freeze_panes = "A4"
    set_col_widths(ws_sum, {1: 26, 2: 18, 3: 12, 4: 20})

    ws_fun = wb.create_sheet("Funnel")
    headers = list(df_funnel.columns)
    for c, h in enumerate(headers, start=1):
        ws_fun.cell(row=1, column=c, value=h)
    style_header_row(ws_fun, 1)

    for r_idx, row in enumerate(df_funnel.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws_fun.cell(row=r_idx, column=c_idx, value=val)

    ws_fun.freeze_panes = "A2"
    col_map = {name: i + 1 for i, name in enumerate(headers)}

    def fmt_col(col_name, number_format):
        if col_name not in col_map:
            return
        col = col_map[col_name]
        for rr in range(2, 2 + len(df_funnel)):
            ws_fun.cell(row=rr, column=col).number_format = number_format

    fmt_col("Patients", "0")
    fmt_col("CAC ($/pt)", "$#,##0")
    fmt_col("Stage CAC ($)", "$#,##0")
    fmt_col("Cumulative CAC ($)", "$#,##0")
    fmt_col("TAM Net Ratio", "0.00%")
    fmt_col("SAM Net Ratio", "0.00%")
    fmt_col("Net Activation Ratio", "0.00%")
    set_col_widths(ws_fun, {1: 5, 2: 52, 3: 22, 4: 12, 5: 14, 6: 12, 7: 15, 8: 18, 9: 14, 10: 14, 11: 18})

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_simple_excel(df, sheet_name="Data"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for i, col in enumerate(df.columns, start=1):
        max_len = max(len(str(col)), *(len(str(v)) for v in df[col].head(100).tolist())) if len(df) > 0 else len(str(col))
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 12), 32)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# -----------------------------
# Session state bootstrap
# -----------------------------
def init_session():
    if "models" not in st.session_state:
        st.session_state["models"] = [copy.deepcopy(SPONSOR_DEFAULTS)]
        st.session_state["model_names"] = ["Model 1"]
        st.session_state["active_model_idx"] = 0

init_session()

# -----------------------------
# Page title
# -----------------------------
st.set_page_config(page_title="PharmaROI Calculator", page_icon="", layout="wide")
st.title("PharmaROI Calculator")
st.caption("Build multiple ROI models side-by-side and compare them in the Comparison tab.")

# -----------------------------
# Model management bar
# -----------------------------
mgmt_col1, mgmt_col2, mgmt_col3, mgmt_col4 = st.columns([2, 2, 2, 4])

with mgmt_col1:
    if st.button("➕ Add New Model", use_container_width=True):
        n = len(st.session_state["models"]) + 1
        st.session_state["models"].append(copy.deepcopy(SPONSOR_DEFAULTS))
        st.session_state["model_names"].append(f"Model {n}")
        st.session_state["active_model_idx"] = len(st.session_state["models"]) - 1
        st.rerun()

with mgmt_col2:
    copy_options = st.session_state["model_names"]
    copy_source = st.selectbox(
        "Copy from:",
        options=range(len(copy_options)),
        format_func=lambda i: copy_options[i],
        index=st.session_state["active_model_idx"],
        key="copy_source_select",
        label_visibility="collapsed",
    )
    if st.button("📋 Copy This Model", use_container_width=True):
        source_idx = copy_source
        new_state = copy.deepcopy(st.session_state["models"][source_idx])
        new_name = st.session_state["model_names"][source_idx] + " (copy)"
        st.session_state["models"].append(new_state)
        st.session_state["model_names"].append(new_name)
        st.session_state["active_model_idx"] = len(st.session_state["models"]) - 1
        st.rerun()

with mgmt_col3:
    can_delete = len(st.session_state["models"]) > 1

    if "confirm_delete" not in st.session_state:
        st.session_state["confirm_delete"] = False

    if not st.session_state["confirm_delete"]:
        if st.button("Delete Current", use_container_width=True, disabled=not can_delete):
            st.session_state["confirm_delete"] = True
            st.rerun()
    else:
        idx = st.session_state["active_model_idx"]
        st.warning(f"Delete '{st.session_state['model_names'][idx]}'?")
        confirm_cols = st.columns(2)
        with confirm_cols[0]:
            if st.button("Yes, Delete", use_container_width=True, type="primary"):
                st.session_state["models"].pop(idx)
                st.session_state["model_names"].pop(idx)
                st.session_state["active_model_idx"] = max(0, idx - 1)
                st.session_state["confirm_delete"] = False
                st.rerun()
        with confirm_cols[1]:
            if st.button("Cancel", use_container_width=True):
                st.session_state["confirm_delete"] = False
                st.rerun()

with mgmt_col4:
    idx = st.session_state["active_model_idx"]
    new_name = st.text_input(
        "Rename current model:",
        value=st.session_state["model_names"][idx],
        key=f"rename_model_{idx}",
        label_visibility="collapsed",
        placeholder="Rename current model…",
    )
    if new_name != st.session_state["model_names"][idx]:
        st.session_state["model_names"][idx] = new_name

# -----------------------------
# Tabs
# -----------------------------
tab_labels = st.session_state["model_names"] + ["Comparison"]
tabs = st.tabs(tab_labels)

for model_idx, model_tab in enumerate(tabs[:-1]):
    with model_tab:
        state = st.session_state["models"][model_idx]
        model_name = st.session_state["model_names"][model_idx]
        tab_color = TAB_PALETTE[model_idx % len(TAB_PALETTE)]

        with st.expander("Model Settings", expanded=(model_idx == st.session_state["active_model_idx"])):
            st.session_state["active_model_idx"] = model_idx

            col_r1, col_r2 = st.columns(2)
            with col_r1:
                if st.button("Reset: Sponsor Example", key=f"reset_sponsor_{model_idx}"):
                    st.session_state["models"][model_idx] = copy.deepcopy(SPONSOR_DEFAULTS)
                    st.rerun()
            with col_r2:
                if st.button("Reset: Zero", key=f"reset_zero_{model_idx}"):
                    st.session_state["models"][model_idx] = copy.deepcopy(ZERO_SAMPLE)
                    st.rerun()

            st.markdown("**Base Population**")
            state["base_population"] = st.number_input(
                "Stage 1 — Total Addressable Market",
                min_value=0,
                step=100_000,
                value=int(state["base_population"]),
                key=f"base_pop_{model_idx}",
            )

            st.markdown("**Revenue & Costs**")
            c1, c3 = st.columns(2)
            with c1:
                state["arpp"] = st.number_input(
                    "ARPP ($/year)",
                    min_value=0.0,
                    step=1_000.0,
                    value=float(state["arpp"]),
                    key=f"arpp_{model_idx}",
                )
            with c3:
                state["discount"] = st.slider(
                    "Discount (gross→net)",
                    min_value=0.0,
                    max_value=1.0,
                    step=0.01,
                    value=float(state["discount"]),
                    key=f"discount_{model_idx}",
                )

            st.markdown("**Funnel Stages**")
            stage_names = state.get("stage_names", STAGE_NAMES[:])

            with st.expander("Customize Stage Names"):
                for sidx in range(len(STAGE_NAMES)):
                    stage_names[sidx] = st.text_input(
                        f"Stage {sidx+1} name:",
                        value=stage_names[sidx],
                        key=f"sname_{model_idx}_{sidx}",
                    )
                state["stage_names"] = stage_names

            for sidx, sname in enumerate(stage_names):
                with st.expander(f"{sidx+1}. {sname}", expanded=False):
                    state["stage_active"][sidx] = st.checkbox(
                        "Use this stage",
                        value=bool(state["stage_active"][sidx]),
                        key=f"active_{model_idx}_{sidx}",
                    )
                    if sidx == 0:
                        st.info("Stage 1 is the base population. No ratio applied.")
                    else:
                        disabled = not state["stage_active"][sidx]
                        state["ratios"][sidx] = st.slider(
                            "Funnel ratio",
                            min_value=0.0,
                            max_value=1.0,
                            step=0.01,
                            value=float(state["ratios"][sidx]),
                            disabled=disabled,
                            key=f"ratio_{model_idx}_{sidx}",
                        )
                    if sidx <= 5:
                        disabled = not state["stage_active"][sidx]
                        state["cac"][sidx] = st.number_input(
                            "CAC ($ per patient)",
                            min_value=0.0,
                            step=1.0,
                            value=float(state["cac"][sidx]),
                            disabled=disabled,
                            key=f"cac_{model_idx}_{sidx}",
                        )
                    else:
                        st.caption("CAC auto-calculated from Stage 6")

            st.markdown("**Platform Costs**")
            if "platform_costs" not in state:
                state["platform_costs"] = SPONSOR_DEFAULTS["platform_costs"].copy()
            pc = state["platform_costs"]
            pc_col1, pc_col2 = st.columns(2)
            with pc_col1:
                pc["dario_connect_config"] = st.number_input(
                    "Dario Connect Configuration",
                    min_value=0.0,
                    step=10_000.0,
                    value=float(pc["dario_connect_config"]),
                    key=f"dcc_{model_idx}",
                )
                pc["dario_care_config"] = st.number_input(
                    "Dario Care Configuration",
                    min_value=0.0,
                    step=10_000.0,
                    value=float(pc["dario_care_config"]),
                    key=f"dcarec_{model_idx}",
                )
                pc["sub_dario_connect"] = st.number_input(
                    "Subscription — Dario Connect",
                    min_value=0.0,
                    step=10_000.0,
                    value=float(pc["sub_dario_connect"]),
                    key=f"sdc_{model_idx}",
                )
            with pc_col2:
                pc["sub_dario_care"] = st.number_input(
                    "Subscription — Dario Care",
                    min_value=0.0,
                    step=10_000.0,
                    value=float(pc["sub_dario_care"]),
                    key=f"sdcare_{model_idx}",
                )
                pc["maintenance_support"] = st.number_input(
                    "Maintenance & Support",
                    min_value=0.0,
                    step=10_000.0,
                    value=float(pc["maintenance_support"]),
                    key=f"ms_{model_idx}",
                )
            st.caption(f"Total Platform Costs: {money(sum(pc.values()))}")

            st.markdown("**Optimization ROI Modeling**")
            phased_enabled = st.checkbox(
                "Enable Optimization ROI Modeling",
                value=state.get("phased_enabled", False),
                key=f"phased_enabled_{model_idx}",
            )
            state["phased_enabled"] = phased_enabled

            if phased_enabled:
                st.caption("Set revenue efficiency per optimization phase (100% = full potential).")
                ph_col1, ph_col2, ph_col3 = st.columns(3)
                with ph_col1:
                    eff_0_3_pct = st.slider(
                        "Months 0-3 efficiency",
                        min_value=0,
                        max_value=100,
                        step=1,
                        value=int(state.get("phased_eff_0_3", 0.33) * 100),
                        format="%d%%",
                        key=f"eff_0_3_{model_idx}",
                    )
                    eff_0_3 = eff_0_3_pct / 100
                with ph_col2:
                    eff_3_6_pct = st.slider(
                        "Months 3-6 efficiency",
                        min_value=0,
                        max_value=100,
                        step=1,
                        value=int(state.get("phased_eff_3_6", 0.66) * 100),
                        format="%d%%",
                        key=f"eff_3_6_{model_idx}",
                    )
                    eff_3_6 = eff_3_6_pct / 100
                with ph_col3:
                    eff_6_plus_pct = st.slider(
                        "Months 6+ efficiency",
                        min_value=0,
                        max_value=100,
                        step=1,
                        value=int(state.get("phased_eff_6_plus", 1.0) * 100),
                        format="%d%%",
                        key=f"eff_6_plus_{model_idx}",
                    )
                    eff_6_plus = eff_6_plus_pct / 100

                state["phased_eff_0_3"] = eff_0_3
                state["phased_eff_3_6"] = eff_3_6
                state["phased_eff_6_plus"] = eff_6_plus

            else:
                state["phased_eff_0_3"] = state.get("phased_eff_0_3", 0.33)
                state["phased_eff_3_6"] = state.get("phased_eff_3_6", 0.66)
                state["phased_eff_6_plus"] = state.get("phased_eff_6_plus", 1.0)

        funnel_results, fin = run_model(state)
        _eff_0_3 = state.get("phased_eff_0_3", 0.33)
        _eff_3_6 = state.get("phased_eff_3_6", 0.66)
        _eff_6_plus = state.get("phased_eff_6_plus", 1.0)
        phase_df = build_phase_optimization_df(fin, state, _eff_0_3, _eff_3_6, _eff_6_plus) if pd is not None else None

        tam_patients = funnel_results[0].patients
        sam_patients = funnel_results[1].patients
        activation_patients = funnel_results[5].patients

        st.markdown(
            f"<div style='border-left: 4px solid {tab_color}; padding-left: 12px; margin-bottom: 8px;'><strong style='font-size:1.1rem'>{model_name}</strong></div>",
            unsafe_allow_html=True,
        )

        roi = fin["roi_net"]
        total_cost = fin["funnel_cac_total"] + fin["platform_costs_total"]

        k1, k2, k3, k4, k5, k6 = st.columns(6)
        k1.metric("ROI (Net)", roix(roi) if roi == roi else "—")
        k2.metric("Treated Patients", number(fin["treated_patients"]))
        k3.metric("Net Revenue", money(fin["net_revenue"]))
        k4.metric("Funnel CAC", money(fin["funnel_cac_total"]))
        k5.metric("Total Cost", money(total_cost))
        k6.metric("Net Profit", money(fin["net_profit"]))

        st.markdown(
            f"Gross: **\\${fin['gross_revenue']:,.0f}**  |  "
            f"Discount: **{fin['discount']*100:.1f}%**  |  "
            f"Discount Amount: **\\${fin['gross_revenue'] - fin['net_revenue']:,.0f}**  |  "
            f"Net Revenue per Rx: **\\${(float(state['arpp']) * (1 - fin['discount'])):,.0f}**"
        )

        if state.get("phased_enabled", False):
            st.markdown("### Optimization Phase Summary")
            ph1, ph2, ph3 = st.columns(3)
            phase_lookup = {row["Phase"]: row for _, row in phase_df.iterrows()} if phase_df is not None else {}

            ph1.metric(
                "ROI — Months 0-3",
                roix(phase_lookup.get("Months 0-3", {}).get("ROI", 0.0)),
                delta=f"{phase_lookup.get('Months 0-3', {}).get('Efficiency', 1.0):.0%} efficiency",
                delta_color="off",
            )
            ph2.metric(
                "ROI — Months 3-6",
                roix(phase_lookup.get("Months 3-6", {}).get("ROI", 0.0)),
                delta=f"{phase_lookup.get('Months 3-6', {}).get('Efficiency', 1.0):.0%} efficiency",
                delta_color="off",
            )
            ph3.metric(
                "ROI — Months 6+",
                roix(phase_lookup.get("Months 6-12", {}).get("ROI", 0.0)),
                delta=f"{phase_lookup.get('Months 6-12', {}).get('Efficiency', 1.0):.0%} efficiency",
                delta_color="off",
            )
            st.caption("Optimization phases use the phase-efficiency inputs to show how ROI and net revenue ramp toward full potential.")
        else:
            st.caption("Optimization ROI Modeling is currently disabled. The phase charts below display full-potential values across all phases.")

        st.subheader("Funnel Table")
        table_rows = []
        for ridx, r in enumerate(funnel_results):
            tam_ratio = r.patients / tam_patients if tam_patients > 0 else 0.0
            sam_ratio = r.patients / sam_patients if sam_patients > 0 else 0.0
            net_activation = r.patients / activation_patients if activation_patients > 0 else 0.0

            table_rows.append({
                "#": ridx + 1,
                "Stage": r.name,
                "Status": "Active" if r.active else "Inactive (pass-through)",
                "Ratio Used": "—" if ridx == 0 else pct(r.ratio_used),
                "Patients": float(r.patients),
                "CAC ($/pt)": float(r.cac_per_patient),
                "Stage CAC ($)": float(r.stage_cac),
                "Cumulative CAC ($)": float(r.cumulative_cac),
                "TAM Net Ratio": float(tam_ratio),
                "SAM Net Ratio": float(sam_ratio),
                "Net Activation Ratio": float(net_activation),
            })

        if pd is not None:
            df_funnel = pd.DataFrame(table_rows)
            df_display = df_funnel.copy()
            df_display["Patients"] = df_display["Patients"].map(lambda x: f"{x:,.0f}")
            df_display["CAC ($/pt)"] = df_display["CAC ($/pt)"].map(lambda x: f"${x:,.0f}")
            df_display["Stage CAC ($)"] = df_display["Stage CAC ($)"].map(lambda x: f"${x:,.0f}")
            df_display["Cumulative CAC ($)"] = df_display["Cumulative CAC ($)"].map(lambda x: f"${x:,.0f}")
            df_display["TAM Net Ratio"] = df_display["TAM Net Ratio"].map(lambda x: "—" if x > 1.0 else f"{x*100:.2f}%")
            df_display["SAM Net Ratio"] = df_display["SAM Net Ratio"].map(lambda x: "—" if x > 1.0 else f"{x*100:.2f}%")
            df_display["Net Activation Ratio"] = df_display["Net Activation Ratio"].map(lambda x: "—" if x > 1.0 else f"{x*100:.2f}%")
            st.dataframe(df_display, use_container_width=True, hide_index=True)

            st.markdown("### Export")
            ec1, ec2 = st.columns(2)
            with ec1:
                xlsx_bytes = build_polished_excel_report(df_funnel, fin, COLORS)
                st.download_button(
                    "⬇️ Download Excel Report",
                    data=xlsx_bytes,
                    file_name=f"{model_name.replace(' ', '_')}_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_xlsx_{model_idx}",
                )
            with ec2:
                csv_data = df_funnel.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "⬇️ Download CSV",
                    data=csv_data,
                    file_name=f"{model_name.replace(' ', '_')}_funnel.csv",
                    mime="text/csv",
                    key=f"dl_csv_{model_idx}",
                )
        else:
            st.write(table_rows)

        st.subheader("Visuals")
        chart_row1_col1, chart_row1_col2 = st.columns(2)

        with chart_row1_col1:
            st.plotly_chart(plotly_waterfall(fin), use_container_width=True, key=f"waterfall_{model_idx}")

        with chart_row1_col2:
            if pd is not None and phase_df is not None:
                st.plotly_chart(
                    plotly_phase_step_chart(
                        phase_df,
                        y_col="ROI",
                        title="Optimization ROI by Phase",
                        y_title="ROI (x)",
                        line_color=tab_color,
                    ),
                    use_container_width=True,
                    key=f"phase_roi_{model_idx}"
                )
            else:
                st.info("Optimization phase chart requires pandas.")

        chart_row2_col1, chart_row2_col2 = st.columns(2)

        with chart_row2_col1:
            if pd is not None and phase_df is not None:
                st.plotly_chart(
                    plotly_phase_step_chart(
                        phase_df,
                        y_col="Net Revenue",
                        title="Optimization Net Revenue by Phase",
                        y_title="Net Revenue",
                        line_color=COLORS["profit"],
                    ),
                    use_container_width=True,
                    key=f"phase_revenue_{model_idx}"
                )
            else:
                st.info("Optimization phase chart requires pandas.")

        with chart_row2_col2:
            if pd is not None:
                funnel_chart_df = pd.DataFrame([{"Stage": r.name, "Patients": r.patients} for r in funnel_results])
                st.plotly_chart(plotly_funnel_patients(funnel_chart_df, tab_color), use_container_width=True, key=f"funnel_patients_{model_idx}")
            else:
                st.info("Funnel visualization requires pandas.")

# -----------------------------
# Comparison Tab
# -----------------------------
with tabs[-1]:
    st.subheader("Model Comparison")

    if len(st.session_state["models"]) < 2:
        st.info("Add at least 2 models to compare them here.")
    else:
        st.markdown("**Select models to compare:**")
        selected_model_names = st.multiselect(
            "Choose models",
            options=st.session_state["model_names"],
            default=st.session_state["model_names"],
            key="comparison_model_select",
            label_visibility="collapsed",
        )

        if len(selected_model_names) < 2:
            st.warning("Please select at least 2 models to compare.")
            st.stop()

        selected_indices = [i for i, name in enumerate(st.session_state["model_names"]) if name in selected_model_names]
        selected_models = [st.session_state["models"][i] for i in selected_indices]
        selected_names = [st.session_state["model_names"][i] for i in selected_indices]

        comparison_rows = []
        phase_rows = []

        for mstate, mname in zip(selected_models, selected_names):
            _, fin = run_model(mstate)
            roi = fin["roi_net"]
            total_cost = fin["funnel_cac_total"] + fin["platform_costs_total"]
            treated = fin["treated_patients"]

            funnel_cac_per_patient = (fin["funnel_cac_total"] / treated) if treated > 0 else 0.0
            platform_costs_per_patient = (fin["platform_costs_total"] / treated) if treated > 0 else 0.0
            total_cost_per_patient = (total_cost / treated) if treated > 0 else 0.0

            comparison_rows.append({
                "Model": mname,
                "Treated Patients": treated,
                "Gross Revenue": fin["gross_revenue"],
                "Net Revenue": fin["net_revenue"],
                "Funnel CAC": fin["funnel_cac_total"],
                "Platform Costs": fin["platform_costs_total"],
                "Total Cost": total_cost,
                "Net Profit": fin["net_profit"],
                "Discount": fin["discount"],
                "ARPP": float(mstate["arpp"]),
                "ROI (Net)": roi if roi == roi else 0.0,
                "Funnel CAC per Treated Patient": funnel_cac_per_patient,
                "Platform Costs per Treated Patient": platform_costs_per_patient,
                "Total Cost per Treated Patient": total_cost_per_patient,
            })

            if pd is not None:
                model_phase_df = build_phase_optimization_df(fin, mstate)
                if model_phase_df is not None:
                    model_phase_df = model_phase_df.copy()
                    model_phase_df["Model"] = mname
                    phase_rows.append(model_phase_df)

        if pd is not None:
            comp_df = pd.DataFrame(comparison_rows)
            color_map = {name: TAB_PALETTE[i % len(TAB_PALETTE)] for i, name in enumerate(selected_names)}

            st.markdown("### Key Metrics")
            disp = comp_df.copy()
            disp["Treated Patients"] = disp["Treated Patients"].map(lambda x: f"{x:,.0f}")
            disp["Gross Revenue"] = disp["Gross Revenue"].map(lambda x: f"${x:,.0f}")
            disp["Net Revenue"] = disp["Net Revenue"].map(lambda x: f"${x:,.0f}")
            disp["Funnel CAC"] = disp["Funnel CAC"].map(lambda x: f"${x:,.0f}")
            disp["Platform Costs"] = disp["Platform Costs"].map(lambda x: f"${x:,.0f}")
            disp["Total Cost"] = disp["Total Cost"].map(lambda x: f"${x:,.0f}")
            disp["Net Profit"] = disp["Net Profit"].map(lambda x: f"${x:,.0f}")
            disp["Discount"] = disp["Discount"].map(lambda x: f"{x*100:.1f}%")
            disp["ARPP"] = disp["ARPP"].map(lambda x: f"${x:,.0f}")
            disp["ROI (Net)"] = disp["ROI (Net)"].map(lambda x: f"{x:.2f}x")
            disp["Funnel CAC per Treated Patient"] = disp["Funnel CAC per Treated Patient"].map(lambda x: f"${x:,.0f}")
            disp["Platform Costs per Treated Patient"] = disp["Platform Costs per Treated Patient"].map(lambda x: f"${x:,.0f}")
            disp["Total Cost per Treated Patient"] = disp["Total Cost per Treated Patient"].map(lambda x: f"${x:,.0f}")
            st.dataframe(disp, use_container_width=True, hide_index=True)

            st.markdown("### Charts")
            chart_col1, chart_col2 = st.columns(2)

            with chart_col1:
                st.plotly_chart(
                    plotly_comparison_bar(
                        comp_df,
                        "ROI (Net)",
                        "Net ROI by Scenario",
                        "ROI (x)",
                        color_map,
                    ),
                    use_container_width=True,
                )

            with chart_col2:
                st.plotly_chart(
                    plotly_comparison_bar(
                        comp_df,
                        "Net Profit",
                        "Net Profit by Scenario",
                        "USD",
                        color_map,
                    ),
                    use_container_width=True,
                )

            chart_col3, chart_col4 = st.columns(2)

            with chart_col3:
                st.plotly_chart(
                    plotly_comparison_bar(
                        comp_df,
                        "Treated Patients",
                        "Treated Patients by Scenario",
                        "Patients",
                        color_map,
                    ),
                    use_container_width=True,
                )
                treated_patients_export = comp_df[["Model", "Treated Patients"]].copy()
                st.download_button(
                    "Download Treated Patients Data (Excel)",
                    data=build_simple_excel(treated_patients_export, "Treated Patients"),
                    file_name="treated_patients_by_scenario.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_treated_patients_by_scenario",
                )

            with chart_col4:
                st.plotly_chart(
                    plotly_comparison_bar(
                        comp_df,
                        "Total Cost",
                        "Total Investment by Scenario",
                        "USD",
                        color_map,
                    ),
                    use_container_width=True,
                )

            st.markdown("### Per-Patient Cost Comparison")
            st.caption("Compares acquisition and platform investment on a per-treated-patient basis across selected scenarios.")

            per_patient_cost_df = pd.DataFrame([
                {
                    "Model": row["Model"],
                    "Metric": "Funnel CAC per Treated Patient",
                    "Cost per Treated Patient": row["Funnel CAC per Treated Patient"],
                }
                for _, row in comp_df.iterrows()
            ] + [
                {
                    "Model": row["Model"],
                    "Metric": "Platform Costs per Treated Patient",
                    "Cost per Treated Patient": row["Platform Costs per Treated Patient"],
                }
                for _, row in comp_df.iterrows()
            ] + [
                {
                    "Model": row["Model"],
                    "Metric": "Total Cost per Treated Patient",
                    "Cost per Treated Patient": row["Total Cost per Treated Patient"],
                }
                for _, row in comp_df.iterrows()
            ])

            st.plotly_chart(plotly_per_patient_costs(per_patient_cost_df, color_map), use_container_width=True)

            per_patient_export_df = comp_df[[
                "Model",
                "Treated Patients",
                "Funnel CAC per Treated Patient",
                "Platform Costs per Treated Patient",
                "Total Cost per Treated Patient",
            ]].copy()

            st.download_button(
                "Download Per-Patient Cost Data (Excel)",
                data=build_simple_excel(per_patient_export_df, "Per Patient Costs"),
                file_name="per_patient_cost_comparison.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_per_patient_cost_comparison",
            )

            if phase_rows:
                st.markdown("### Optimization Comparison by Phase")
                phase_comp_df = pd.concat(phase_rows, ignore_index=True)

                op1, op2 = st.columns(2)
                with op1:
                    st.plotly_chart(
                        plotly_phase_comparison_chart(
                            phase_comp_df,
                            y_col="ROI",
                            title="Optimization ROI by Phase Across Scenarios",
                            y_title="ROI (x)",
                            color_map=color_map,
                        ),
                        use_container_width=True,
                    )

                with op2:
                    st.plotly_chart(
                        plotly_phase_comparison_chart(
                            phase_comp_df,
                            y_col="Net Revenue",
                            title="Optimization Net Revenue by Phase Across Scenarios",
                            y_title="Net Revenue",
                            color_map=color_map,
                        ),
                        use_container_width=True,
                    )

                optimization_export_df = phase_comp_df[[
                    "Model",
                    "Phase",
                    "Efficiency",
                    "ROI",
                    "Net Revenue",
                ]].copy()

                st.download_button(
                    "Download Optimization Phase Data (Excel)",
                    data=build_simple_excel(optimization_export_df, "Optimization Phases"),
                    file_name="optimization_phase_comparison.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_optimization_phase_comparison",
                )

            st.markdown("### Model Diff View")
            if len(selected_names) >= 2:
                diff_col1, diff_col2 = st.columns(2)
                with diff_col1:
                    diff_model_a = st.selectbox("Model A:", options=selected_names, index=0, key="diff_model_a")
                with diff_col2:
                    remaining = [n for n in selected_names if n != diff_model_a]
                    diff_model_b = st.selectbox("Model B:", options=remaining, index=0, key="diff_model_b")

                idx_a = st.session_state["model_names"].index(diff_model_a)
                idx_b = st.session_state["model_names"].index(diff_model_b)
                state_a = st.session_state["models"][idx_a]
                state_b = st.session_state["models"][idx_b]

                diff_rows = []

                top_params = [
                    ("Base Population", "base_population", "{:,.0f}"),
                    ("ARPP", "arpp", "${:,.0f}"),
                    ("Treatment Years", "treatment_years", "{:.1f}"),
                    ("Discount", "discount", "{:.1%}"),
                ]
                for label, key, fmt in top_params:
                    val_a = state_a.get(key, 0)
                    val_b = state_b.get(key, 0)
                    if val_a != val_b:
                        diff_rows.append({
                            "Parameter": label,
                            f"{diff_model_a}": fmt.format(val_a),
                            f"{diff_model_b}": fmt.format(val_b),
                            "Difference": fmt.format(val_b - val_a) if "%" not in fmt else f"{(val_b - val_a) * 100:+.1f}pp",
                        })

                for sidx in range(len(STAGE_NAMES)):
                    ratio_a = state_a["ratios"][sidx]
                    ratio_b = state_b["ratios"][sidx]
                    if ratio_a != ratio_b and sidx > 0:
                        diff_rows.append({
                            "Parameter": f"Stage {sidx+1} Ratio",
                            f"{diff_model_a}": f"{ratio_a:.1%}",
                            f"{diff_model_b}": f"{ratio_b:.1%}",
                            "Difference": f"{(ratio_b - ratio_a) * 100:+.1f}pp",
                        })

                    cac_a = state_a["cac"][sidx]
                    cac_b = state_b["cac"][sidx]
                    if cac_a != cac_b:
                        diff_rows.append({
                            "Parameter": f"Stage {sidx+1} CAC",
                            f"{diff_model_a}": f"${cac_a:,.0f}",
                            f"{diff_model_b}": f"${cac_b:,.0f}",
                            "Difference": f"${cac_b - cac_a:+,.0f}",
                        })

                if diff_rows:
                    diff_df = pd.DataFrame(diff_rows)
                    st.dataframe(diff_df, use_container_width=True, hide_index=True)
                else:
                    st.success("These two models have identical parameters!")
            else:
                st.info("Select at least 2 models above to see a diff view.")

            st.markdown("### Export Comparison")
            comp_csv = comp_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                "⬇️ Download Comparison CSV",
                data=comp_csv,
                file_name="pharmaroi_comparison.csv",
                mime="text/csv",
            )

st.divider()
st.subheader("How to interpret")
st.write("""
- Each **model tab** is fully independent — tweak funnel stages, ratios, CAC, ARPP, and discount separately.
- Use **Add New Model** or **Duplicate Current** to create variants.
- The **Comparison** tab shows all models side-by-side with charts and a downloadable table.
- **ROI (Net)** = Net Revenue / (Funnel CAC + Platform Costs)
- **Net Profit** = Net Revenue − Funnel CAC − Platform Costs
- **Net Revenue** = Gross Revenue × (1 − Discount)
- **TAM Net Ratio** = Patients at Stage / Stage 1 (Total Addressable Market)
- **SAM Net Ratio** = Patients at Stage / Stage 2 (F2 and F3)
- **Net Activation Ratio** = Patients at Stage / Stage 6 (Activation onto Dario Connect)

**Optimization phase assumption:**
- Optimization phases use the selected efficiency values for Months 0–3, Months 3–6, and Months 6+.
- Phase ROI is shown as full-potential ROI scaled by the phase efficiency.
- Phase Net Revenue is shown as full-potential Net Revenue scaled by the phase efficiency.
- These optimization charts are a visualization layer and do not change the core annual model logic.
""")